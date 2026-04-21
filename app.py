import os
import re
import uuid
import secrets
from functools import wraps
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory, session, Response
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3

APP_VERSION = '1.5.9'

app = Flask(__name__)

# Data directory — configurable via env for Docker volume mounting
DATA_DIR = os.environ.get('WM_DATA_DIR', os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Secret key: env var > file > auto-generate
if os.environ.get('SECRET_KEY'):
    app.config['SECRET_KEY'] = os.environ['SECRET_KEY']
else:
    SECRET_KEY_FILE = os.path.join(DATA_DIR, '.secret_key')
    if os.path.exists(SECRET_KEY_FILE):
        with open(SECRET_KEY_FILE, 'r') as f:
            app.config['SECRET_KEY'] = f.read().strip()
    else:
        key = secrets.token_hex(32)
        with open(SECRET_KEY_FILE, 'w') as f:
            f.write(key)
        os.chmod(SECRET_KEY_FILE, 0o600)
        app.config['SECRET_KEY'] = key

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Set secure flag when the deployment is behind HTTPS. Toggle via env var so
# dev-over-HTTP still works. WM_SECURE_COOKIES=1 flips it on in production.
_SECURE_COOKIES = os.environ.get('WM_SECURE_COOKIES', '').lower() in ('1', 'true', 'yes')
app.config['SESSION_COOKIE_SECURE'] = _SECURE_COOKIES
app.config['REMEMBER_COOKIE_SECURE'] = _SECURE_COOKIES
# 90 days — keep users signed in long-term. Both values matter:
#   PERMANENT_SESSION_LIFETIME  → the Flask session cookie
#   REMEMBER_COOKIE_DURATION    → Flask-Login's "remember me" cookie that re-authenticates
app.config['PERMANENT_SESSION_LIFETIME'] = 60 * 60 * 24 * 90  # 90 days
app.config['REMEMBER_COOKIE_DURATION'] = 60 * 60 * 24 * 90  # 90 days
app.config['REMEMBER_COOKIE_HTTPONLY'] = True
app.config['REMEMBER_COOKIE_SAMESITE'] = 'Lax'

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# Account-lockout policy: lock after N consecutive failures for LOCKOUT_MINUTES.
LOGIN_FAIL_LIMIT = 5
LOCKOUT_MINUTES = 15
PASSWORD_MIN_LENGTH = 8


@app.after_request
def _security_headers(resp):
    """Add defense-in-depth headers on every response."""
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'DENY')
    resp.headers.setdefault('Referrer-Policy', 'same-origin')
    resp.headers.setdefault('Permissions-Policy',
                            'geolocation=(), camera=(), microphone=(), payment=()')
    if _SECURE_COOKIES:
        resp.headers.setdefault('Strict-Transport-Security',
                                'max-age=31536000; includeSubDomains')
    return resp

DATABASE = os.path.join(DATA_DIR, 'warehouse.db')

# ══════════════════════════════════════════
#  LOGIN MANAGER
# ══════════════════════════════════════════

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = ''


class User(UserMixin):
    def __init__(self, id, username, display_name, role, active):
        self.id = id
        self.username = username
        self.display_name = display_name
        self.role = role
        self.is_active_user = active

    def get_id(self):
        return str(self.id)

    @property
    def is_admin(self):
        return self.role == 'admin'

    @property
    def can_edit(self):
        return self.role in ('admin', 'editor', 'supervisor')

    @property
    def can_view_audit(self):
        return self.role in ('admin', 'supervisor')

    def is_active(self):
        return bool(self.is_active_user)


@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    if row:
        return User(row['id'], row['username'], row['display_name'], row['role'], row['active'])
    return None


def admin_required(f):
    """Decorator: requires admin role."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated


def editor_required(f):
    """Decorator: requires admin, editor, or supervisor role."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.can_edit:
            return jsonify({'error': 'Editor access required'}), 403
        return f(*args, **kwargs)
    return decorated


def audit_view_required(f):
    """Decorator: requires admin or supervisor role (audit trail access)."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.can_view_audit:
            return jsonify({'error': 'Audit access required'}), 403
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════

def _natural_sort_key(s):
    """Return a sort key for natural alphanumeric sorting.
    Each chunk becomes a (type, value) tuple so ints and strings never compare directly."""
    if not s:
        return [(0, '')]
    parts = re.split(r'(\d+)', str(s))
    # (0, str) for text chunks, (1, int) for number chunks — type flag ensures same-type comparison
    return [(1, int(c)) if c.isdigit() else (0, c.lower()) for c in parts if c]


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


# ──────────────────────────────────────────
#  MIGRATION SYSTEM
#
#  How it works:
#    - Each migration is a function that receives a db connection
#    - Migrations are registered in MIGRATIONS as (version, function) tuples
#    - On startup, only NEW migrations (version > current) are executed
#    - The schema_version table tracks what's been applied
#
#  To make a schema change in the future:
#    1. Write a new function, e.g. migrate_v3(conn)
#    2. Put your ALTER TABLE / CREATE INDEX / etc. inside it
#    3. Append it to MIGRATIONS: (3, migrate_v3)
#    4. That's it — existing data is preserved
#
#  Example future migration:
#
#    def migrate_v3(conn):
#        """Add a 'condition' column to parts."""
#        conn.execute("ALTER TABLE parts ADD COLUMN condition TEXT DEFAULT ''")
#
#    Then add to MIGRATIONS list:  (3, migrate_v3),
# ──────────────────────────────────────────

def migrate_v1(conn):
    """Initial schema: users table, parts table, indexes, default admin."""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE COLLATE NOCASE,
            display_name TEXT DEFAULT '',
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user' CHECK(role IN ('admin', 'user')),
            active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS parts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL CHECK(category IN ('engine', 'head', 'transmission')),
            sku TEXT DEFAULT '',
            location TEXT DEFAULT '',
            fitment_vehicle TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            image_filename TEXT DEFAULT '',
            head_engine TEXT DEFAULT '',
            head_part TEXT DEFAULT '',
            foundry_number TEXT DEFAULT '',
            foundry TEXT DEFAULT '',
            head_number TEXT DEFAULT '',
            head_type TEXT DEFAULT '',
            engine_name TEXT DEFAULT '',
            engine_head TEXT DEFAULT '',
            engine_litre TEXT DEFAULT '',
            engine_date_stamp TEXT DEFAULT '',
            engine_turns INTEGER DEFAULT 0,
            trans_gear_condition TEXT DEFAULT '',
            trans_spins INTEGER DEFAULT 0,
            trans_shifts INTEGER DEFAULT 0,
            trans_date_code TEXT DEFAULT '',
            trans_stamped_numbers TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_parts_category ON parts(category);
        CREATE INDEX IF NOT EXISTS idx_parts_sku ON parts(sku);
        CREATE INDEX IF NOT EXISTS idx_parts_fitment ON parts(fitment_vehicle);
    ''')

    # Create default admin if no users exist
    count = conn.execute("SELECT COUNT(*) as n FROM users").fetchone()['n']
    if count == 0:
        pw_hash = generate_password_hash('admin', method='pbkdf2:sha256')
        conn.execute(
            "INSERT INTO users (username, display_name, password_hash, role, active) VALUES (?, ?, ?, ?, ?)",
            ('admin', 'Administrator', pw_hash, 'admin', 1)
        )
        print("\n  ╔══════════════════════════════════════════════╗")
        print("  ║  DEFAULT ADMIN ACCOUNT CREATED               ║")
        print("  ║  Username: admin                              ║")
        print("  ║  Password: admin                              ║")
        print("  ║  ⚠  CHANGE THIS PASSWORD IMMEDIATELY!        ║")
        print("  ╚══════════════════════════════════════════════╝\n")


def migrate_v2(conn):
    """Add location index for sort performance."""
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_location ON parts(location)")


def migrate_v3(conn):
    """Add sold, sold_date, and head_old_number fields."""
    conn.execute("ALTER TABLE parts ADD COLUMN sold INTEGER DEFAULT 0")
    conn.execute("ALTER TABLE parts ADD COLUMN sold_date TEXT DEFAULT ''")
    conn.execute("ALTER TABLE parts ADD COLUMN head_old_number TEXT DEFAULT ''")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_sold ON parts(sold)")


def migrate_v4(conn):
    """Add editor role, rename user role to viewer."""
    conn.executescript('''
        CREATE TABLE users_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE COLLATE NOCASE,
            display_name TEXT DEFAULT '',
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'viewer' CHECK(role IN ('admin', 'editor', 'viewer')),
            active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        INSERT INTO users_new (id, username, display_name, password_hash, role, active, created_at)
            SELECT id, username, display_name, password_hash,
                   CASE WHEN role = 'user' THEN 'viewer' ELSE role END,
                   active, created_at
            FROM users;
        DROP TABLE users;
        ALTER TABLE users_new RENAME TO users;
    ''')


def migrate_v5(conn):
    """Add part_images table for multi-image support, migrate existing images."""
    conn.execute('''
        CREATE TABLE IF NOT EXISTS part_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (part_id) REFERENCES parts(id) ON DELETE CASCADE
        )
    ''')
    conn.execute("CREATE INDEX IF NOT EXISTS idx_part_images_part ON part_images(part_id)")
    # Migrate existing image_filename values into the new table
    rows = conn.execute("SELECT id, image_filename FROM parts WHERE image_filename != '' AND image_filename IS NOT NULL").fetchall()
    for row in rows:
        conn.execute("INSERT INTO part_images (part_id, filename, sort_order) VALUES (?, ?, 0)",
                     (row['id'], row['image_filename']))


def migrate_v6(conn):
    """Convert engine_turns, trans_spins, trans_shifts from integer toggles to text radio values."""
    # engine_turns: 1 -> 'yes', 0 -> 'untested'
    conn.execute("UPDATE parts SET engine_turns = CASE WHEN engine_turns = 1 THEN 'yes' WHEN engine_turns = 0 THEN 'untested' ELSE 'untested' END WHERE category = 'engine'")
    # trans_spins: 0 -> 'untested', 1 -> 'yes'
    conn.execute("UPDATE parts SET trans_spins = CASE WHEN trans_spins = 1 THEN 'yes' WHEN trans_spins = 0 THEN 'untested' ELSE 'untested' END WHERE category = 'transmission'")
    # trans_shifts: 0 -> 'untested', 1 -> 'yes'
    conn.execute("UPDATE parts SET trans_shifts = CASE WHEN trans_shifts = 1 THEN 'yes' WHEN trans_shifts = 0 THEN 'untested' ELSE 'untested' END WHERE category = 'transmission'")


def migrate_v7(conn):
    """Normalize engine_turns values to yes/no/untested."""
    conn.execute("UPDATE parts SET engine_turns = 'yes' WHERE engine_turns = 'turns'")
    conn.execute("UPDATE parts SET engine_turns = 'no' WHERE engine_turns = 'does_not_turn'")


def migrate_v8(conn):
    """Add custom categories and category_fields tables, custom_data column on parts."""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            icon TEXT DEFAULT '',
            color TEXT DEFAULT '#4a8eff',
            sort_order INTEGER DEFAULT 0,
            is_builtin INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS category_fields (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_slug TEXT NOT NULL,
            field_key TEXT NOT NULL,
            field_label TEXT NOT NULL,
            field_type TEXT NOT NULL DEFAULT 'text',
            radio_options TEXT DEFAULT '',
            show_on_card INTEGER DEFAULT 0,
            show_in_table INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (category_slug) REFERENCES categories(slug) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_cf_slug ON category_fields(category_slug);
    ''')
    # Add custom_data column for flexible JSON storage
    try:
        conn.execute("ALTER TABLE parts ADD COLUMN custom_data TEXT DEFAULT '{}'")
    except Exception:
        pass  # column may already exist

    # Register built-in categories
    for slug, name, color, order in [
        ('engine', 'Engine', '#f59e0b', 1),
        ('head', 'Cylinder Head', '#10b981', 2),
        ('transmission', 'Transmission', '#a78bfa', 3),
    ]:
        existing = conn.execute("SELECT id FROM categories WHERE slug = ?", (slug,)).fetchone()
        if not existing:
            conn.execute("INSERT INTO categories (slug, name, color, sort_order, is_builtin) VALUES (?, ?, ?, ?, 1)",
                         (slug, name, color, order))

    # Remove the CHECK constraint on category column by rebuilding parts table
    # This allows custom category slugs
    cols_info = conn.execute("PRAGMA table_info(parts)").fetchall()
    col_names = [c['name'] for c in cols_info]
    cols_str = ', '.join(col_names)

    # Create new table without CHECK constraint on category
    create_cols = []
    for c in cols_info:
        col_def = f"{c['name']} {c['type']}"
        if c['name'] == 'category':
            col_def = "category TEXT NOT NULL"
        elif c['notnull'] and c['dflt_value'] is not None:
            col_def += f" NOT NULL DEFAULT {c['dflt_value']}"
        elif c['notnull']:
            col_def += " NOT NULL"
        elif c['dflt_value'] is not None:
            col_def += f" DEFAULT {c['dflt_value']}"
        if c['pk']:
            col_def += " PRIMARY KEY AUTOINCREMENT"
        create_cols.append(col_def)

    conn.execute(f"CREATE TABLE parts_new ({', '.join(create_cols)})")
    conn.execute(f"INSERT INTO parts_new ({cols_str}) SELECT {cols_str} FROM parts")
    conn.execute("DROP TABLE parts")
    conn.execute("ALTER TABLE parts_new RENAME TO parts")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_category ON parts(category)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_sku ON parts(sku)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_fitment ON parts(fitment_vehicle)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_location ON parts(location)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_sold ON parts(sold)")


def migrate_v9(conn):
    """Add product_number column and assign retroactively starting at WM-1000000."""
    try:
        conn.execute("ALTER TABLE parts ADD COLUMN product_number TEXT DEFAULT ''")
    except Exception:
        pass
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_parts_product_number ON parts(product_number) WHERE product_number != ''")
    # Assign numbers to existing parts ordered by id
    rows = conn.execute("SELECT id FROM parts WHERE product_number = '' OR product_number IS NULL ORDER BY id").fetchall()
    # Find the next available number
    max_row = conn.execute("SELECT product_number FROM parts WHERE product_number LIKE 'WM-%' ORDER BY product_number DESC LIMIT 1").fetchone()
    if max_row and max_row['product_number']:
        try:
            next_num = int(max_row['product_number'].replace('WM-', '')) + 1
        except ValueError:
            next_num = 1000000
    else:
        next_num = 1000000
    for row in rows:
        conn.execute("UPDATE parts SET product_number = ? WHERE id = ?", (f"WM-{next_num}", row['id']))
        next_num += 1


def migrate_v10(conn):
    """Register field definitions for original categories and mark all categories as non-builtin."""
    import json as json_mod

    # Define field sets for the three original categories
    head_fields = [
        ('sku', 'SKU', 'text', '', 0, 1),
        ('head_old_number', 'Old SKU', 'text', '', 0, 1),
        ('location', 'Location', 'text', '', 1, 1),
        ('head_engine', 'Engine', 'text', '', 1, 1),
        ('head_part', 'Part', 'text', '', 0, 0),
        ('foundry_number', 'Foundry Number', 'text', '', 0, 0),
        ('foundry', 'Foundry', 'text', '', 0, 1),
        ('head_number', 'Head Number', 'text', '', 1, 1),
        ('head_type', 'Type', 'text', '', 0, 0),
        ('fitment_vehicle', 'Fitment Vehicle', 'text', '', 1, 1),
        ('sold', 'Sold', 'toggle', '', 0, 1),
        ('sold_date', 'Sold Date', 'text', '', 0, 0),
        ('notes', 'Notes', 'textarea', '', 0, 0),
    ]

    engine_fields = [
        ('sku', 'SKU', 'text', '', 0, 1),
        ('location', 'Location', 'text', '', 1, 1),
        ('engine_turns', 'Turns', 'radio', 'Yes,No,Untested', 0, 1),
        ('engine_name', 'Engine', 'text', '', 1, 1),
        ('engine_head', 'Head', 'text', '', 0, 0),
        ('engine_litre', 'Litre', 'text', '', 1, 1),
        ('engine_date_stamp', 'Date Stamp', 'text', '', 0, 1),
        ('fitment_vehicle', 'Fitment Vehicle', 'text', '', 1, 1),
        ('sold', 'Sold', 'toggle', '', 0, 1),
        ('sold_date', 'Sold Date', 'text', '', 0, 0),
        ('notes', 'Notes', 'textarea', '', 0, 0),
    ]

    trans_fields = [
        ('sku', 'SKU', 'text', '', 0, 1),
        ('location', 'Location', 'text', '', 1, 1),
        ('trans_spins', 'Spins', 'radio', 'Yes,No,Untested', 0, 1),
        ('trans_shifts', 'Shifts', 'radio', 'Yes,No,Untested', 0, 1),
        ('trans_date_code', 'Date Code', 'text', '', 0, 1),
        ('trans_stamped_numbers', 'Stamped Numbers', 'text', '', 0, 1),
        ('fitment_vehicle', 'Fitment Vehicle', 'text', '', 1, 1),
        ('sold', 'Sold', 'toggle', '', 0, 1),
        ('sold_date', 'Sold Date', 'text', '', 0, 0),
        ('notes', 'Notes', 'textarea', '', 0, 0),
    ]

    cat_field_map = {
        'head': head_fields,
        'engine': engine_fields,
        'transmission': trans_fields,
    }

    for slug, fields_def in cat_field_map.items():
        # Only insert fields if none exist yet for this category
        existing = conn.execute("SELECT COUNT(*) as n FROM category_fields WHERE category_slug = ?", (slug,)).fetchone()['n']
        if existing == 0:
            for i, (key, label, ftype, radio_opts, show_card, show_table) in enumerate(fields_def):
                conn.execute(
                    "INSERT INTO category_fields (category_slug, field_key, field_label, field_type, radio_options, show_on_card, show_in_table, sort_order) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (slug, key, label, ftype, radio_opts, show_card, show_table, i)
                )

    # Mark all categories as non-builtin (deletable)
    conn.execute("UPDATE categories SET is_builtin = 0")

    # Migrate existing hardcoded column data into custom_data JSON for all parts
    rows = conn.execute("SELECT * FROM parts").fetchall()
    hardcoded_keys = {
        'head': ['sku', 'location', 'head_engine', 'head_part', 'foundry_number', 'foundry',
                 'head_number', 'head_type', 'head_old_number', 'fitment_vehicle', 'sold', 'sold_date', 'notes'],
        'engine': ['sku', 'location', 'engine_turns', 'engine_name', 'engine_head', 'engine_litre',
                   'engine_date_stamp', 'fitment_vehicle', 'sold', 'sold_date', 'notes'],
        'transmission': ['sku', 'location', 'trans_spins', 'trans_shifts', 'trans_date_code',
                         'trans_stamped_numbers', 'fitment_vehicle', 'sold', 'sold_date', 'notes'],
    }
    # Shared fields that stay in their own columns (not moved to custom_data)
    shared_keys = {'sku', 'location', 'fitment_vehicle', 'sold', 'sold_date', 'notes',
                   'category', 'id', 'image_filename', 'created_at', 'updated_at', 'product_number', 'custom_data'}

    for row in rows:
        r = dict(row)
        cat = r.get('category', '')
        if cat not in hardcoded_keys:
            continue
        existing_cd = {}
        try:
            existing_cd = json_mod.loads(r.get('custom_data', '{}') or '{}')
        except Exception:
            pass
        # Move category-specific hardcoded columns into custom_data
        for key in hardcoded_keys[cat]:
            if key in shared_keys:
                continue
            val = r.get(key, '')
            if val and key not in existing_cd:
                existing_cd[key] = str(val) if val else ''
        if existing_cd:
            conn.execute("UPDATE parts SET custom_data = ? WHERE id = ?",
                         (json_mod.dumps(existing_cd), r['id']))


def migrate_v11(conn):
    """Deduplicate category_fields created by v10 race condition across Gunicorn workers."""
    # For each category, keep only the first set of fields (lowest IDs)
    cats = conn.execute("SELECT DISTINCT category_slug FROM category_fields").fetchall()
    for cat in cats:
        slug = cat['category_slug']
        # Get all fields grouped by field_key, keeping the one with lowest id
        rows = conn.execute(
            "SELECT field_key, MIN(id) as keep_id FROM category_fields WHERE category_slug = ? GROUP BY field_key",
            (slug,)
        ).fetchall()
        keep_ids = [r['keep_id'] for r in rows]
        if keep_ids:
            placeholders = ','.join(['?'] * len(keep_ids))
            conn.execute(
                f"DELETE FROM category_fields WHERE category_slug = ? AND id NOT IN ({placeholders})",
                [slug] + keep_ids
            )


def migrate_v12(conn):
    """Remove dash from product numbers: WM-1000000 becomes WM1000000."""
    conn.execute("UPDATE parts SET product_number = REPLACE(product_number, 'WM-', 'WM') WHERE product_number LIKE 'WM-%'")


def migrate_v13(conn):
    """Add flagged column, update head category fields, add Posted to Web to all categories."""
    # Add flagged column
    try:
        conn.execute("ALTER TABLE parts ADD COLUMN flagged INTEGER DEFAULT 0")
    except Exception:
        pass
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_flagged ON parts(flagged)")

    # Update head category: remove Old SKU and Head Number from table view
    conn.execute("UPDATE category_fields SET show_in_table = 0 WHERE category_slug = 'head' AND field_key = 'head_old_number'")
    conn.execute("UPDATE category_fields SET show_in_table = 0 WHERE category_slug = 'head' AND field_key = 'head_number'")

    # Add "Posted to Web" field to all existing categories if not already present
    cats = conn.execute("SELECT slug FROM categories").fetchall()
    for cat in cats:
        slug = cat['slug']
        existing = conn.execute("SELECT id FROM category_fields WHERE category_slug = ? AND field_key = 'posted_to_web'", (slug,)).fetchone()
        if not existing:
            # Insert before sold field
            sold_order = conn.execute("SELECT sort_order FROM category_fields WHERE category_slug = ? AND field_key = 'sold'", (slug,)).fetchone()
            order = (sold_order['sort_order'] if sold_order else 99) - 1
            conn.execute(
                "INSERT INTO category_fields (category_slug, field_key, field_label, field_type, radio_options, show_on_card, show_in_table, sort_order) VALUES (?, 'posted_to_web', 'Posted to Web', 'toggle', '', 0, 1, ?)",
                (slug, order)
            )


def migrate_v14(conn):
    """Reassign product numbers to 5-digit format starting at WM00001."""
    rows = conn.execute("SELECT id FROM parts ORDER BY id").fetchall()
    for i, row in enumerate(rows):
        num = i + 1
        if num < 100000:
            pn = f"WM{num:05d}"
        else:
            pn = f"WM{num}"
        conn.execute("UPDATE parts SET product_number = ? WHERE id = ?", (pn, row['id']))


def migrate_v15(conn):
    """Add work_orders, work_order_notes, app_settings tables for the work order system."""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS work_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            wo_number TEXT NOT NULL UNIQUE,
            request_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            warehouse_location TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            quote_invoice TEXT DEFAULT '',
            sales_person TEXT DEFAULT '',
            vehicle TEXT DEFAULT '',
            vin TEXT DEFAULT '',
            priority TEXT DEFAULT 'Normal',
            notes TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'requested' CHECK(status IN ('requested','flagged','delivered')),
            flag_note TEXT DEFAULT '',
            created_by TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_wo_status ON work_orders(status);
        CREATE INDEX IF NOT EXISTS idx_wo_request_date ON work_orders(request_date);

        CREATE TABLE IF NOT EXISTS work_order_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_order_id INTEGER NOT NULL,
            note TEXT DEFAULT '',
            author TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (work_order_id) REFERENCES work_orders(id) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_won_wo ON work_order_notes(work_order_id);

        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT DEFAULT ''
        );
    ''')

    import json as json_mod
    defaults = {
        'wo_locations': json_mod.dumps([]),
        'wo_salespeople': json_mod.dumps([]),  # [{name, email}, ...]
        'wo_priorities': json_mod.dumps(['Normal', 'Next Day Air']),
        'smtp_config': json_mod.dumps({
            'host': '', 'port': 587, 'username': '', 'password': '',
            'use_tls': True, 'from_email': '', 'from_name': 'Warehouse Manager'
        }),
    }
    for k, v in defaults.items():
        existing = conn.execute("SELECT key FROM app_settings WHERE key = ?", (k,)).fetchone()
        if not existing:
            conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?)", (k, v))


def migrate_v16(conn):
    """Add parts_json column to work_orders for the repeatable parts-requested list."""
    try:
        conn.execute("ALTER TABLE work_orders ADD COLUMN parts_json TEXT DEFAULT '[]'")
    except Exception:
        pass


def migrate_v23(conn):
    """Track whether a work order has ever been archived.
    Editors cannot delete a previously-archived-then-reopened WO — only admins
    and supervisors can — and the editor UI shows Re-Archive instead."""
    try:
        conn.execute("ALTER TABLE work_orders ADD COLUMN was_archived INTEGER DEFAULT 0")
    except Exception:
        pass
    # Backfill: anything that's already archived has definitely "been archived"
    conn.execute("UPDATE work_orders SET was_archived = 1 WHERE archived_at IS NOT NULL")


def migrate_v22(conn):
    """Delayed archival for delivered work orders.
    archive_after holds a local-time 23:00 threshold set when marked delivered.
    archived_at is set by the sweep (or the Archive Now button) and is what
    separates the Archive view from the Active list."""
    try:
        conn.execute("ALTER TABLE work_orders ADD COLUMN archived_at TIMESTAMP")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE work_orders ADD COLUMN archive_after TIMESTAMP")
    except Exception:
        pass
    # Backfill: any existing delivered WO gets archived immediately so behavior
    # matches the old "delivered == archive" model for pre-upgrade records.
    conn.execute(
        "UPDATE work_orders SET archived_at = COALESCE(completed_at, CURRENT_TIMESTAMP) "
        "WHERE status = 'delivered' AND archived_at IS NULL"
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_wo_archived_at ON work_orders(archived_at)")


def migrate_v21(conn):
    """Add needs_audit + audit_note columns to parts."""
    try:
        conn.execute("ALTER TABLE parts ADD COLUMN needs_audit INTEGER DEFAULT 0")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE parts ADD COLUMN audit_note TEXT DEFAULT ''")
    except Exception:
        pass
    conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_needs_audit ON parts(needs_audit)")


def migrate_v20(conn):
    """Add 'supervisor' to the users role CHECK constraint (editor perms + audit view)."""
    conn.executescript('''
        CREATE TABLE users_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE COLLATE NOCASE,
            display_name TEXT DEFAULT '',
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'viewer' CHECK(role IN ('admin', 'editor', 'supervisor', 'viewer')),
            active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        INSERT INTO users_new (id, username, display_name, password_hash, role, active, created_at)
            SELECT id, username, display_name, password_hash, role, active, created_at FROM users;
        DROP TABLE users;
        ALTER TABLE users_new RENAME TO users;
    ''')


def migrate_v19(conn):
    """Add note_type to work_order_notes so flag notes and general running notes
    can be distinguished. Existing entries are tagged 'flag'."""
    try:
        conn.execute("ALTER TABLE work_order_notes ADD COLUMN note_type TEXT DEFAULT 'flag'")
    except Exception:
        pass


def migrate_v18(conn):
    """Reformat work-order numbers from WO##### to WO-##### for existing records."""
    rows = conn.execute(
        "SELECT id, wo_number FROM work_orders WHERE wo_number LIKE 'WO%' AND wo_number NOT LIKE 'WO-%'"
    ).fetchall()
    for r in rows:
        num_part = r['wo_number'][2:]  # drop 'WO'
        new_num = f"WO-{num_part}"
        conn.execute("UPDATE work_orders SET wo_number = ? WHERE id = ?", (new_num, r['id']))


def migrate_v17(conn):
    """Add work_order_audit table for per-record change history."""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS work_order_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_order_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            actor TEXT DEFAULT '',
            description TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_woa_wo ON work_order_audit(work_order_id);
        CREATE INDEX IF NOT EXISTS idx_woa_created ON work_order_audit(created_at);
    ''')


def migrate_v24(conn):
    """Add work_order_part_photos table for per-part photo uploads on WOs.
    Each part in parts_json gets a stable UUID `key` so photos survive
    re-ordering/edits; the key is generated lazily on first load."""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS work_order_part_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_order_id INTEGER NOT NULL,
            part_key TEXT NOT NULL,
            filename TEXT NOT NULL,
            comment TEXT DEFAULT '',
            uploaded_by TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (work_order_id) REFERENCES work_orders(id) ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_wopp_wo_key ON work_order_part_photos(work_order_id, part_key);
    ''')


def migrate_v29(conn):
    """Add was_delivered + created_by_user_id to work_orders so delete gating
    can distinguish reopened-from-delivered (never delete-able) from
    reopened-from-undelivered-archive (deletable), and so delete can check
    the originator reliably instead of fuzzy-matching display names.
    Backfills was_delivered=1 for anything previously delivered/archived and
    resolves created_by_user_id by display_name/username match."""
    try:
        conn.execute("ALTER TABLE work_orders ADD COLUMN was_delivered INTEGER DEFAULT 0")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE work_orders ADD COLUMN created_by_user_id INTEGER DEFAULT NULL")
    except Exception:
        pass
    # Anything currently delivered, previously completed, or archived was once
    # delivered under the legacy rules (archive used to require delivery).
    conn.execute(
        "UPDATE work_orders SET was_delivered = 1 "
        "WHERE status = 'delivered' OR completed_at IS NOT NULL OR archived_at IS NOT NULL"
    )
    conn.execute("""
        UPDATE work_orders
           SET created_by_user_id = (
               SELECT u.id FROM users u
                WHERE LOWER(u.display_name) = LOWER(work_orders.created_by)
                   OR LOWER(u.username) = LOWER(work_orders.created_by)
                LIMIT 1
           )
         WHERE created_by_user_id IS NULL
           AND created_by IS NOT NULL AND created_by != ''
    """)


def migrate_v28(conn):
    """Account-lockout state on users. After too many failed logins we set
    locked_until; admins can clear it. failed_login_count decays to zero on
    a successful sign-in or after an admin unlock."""
    for col, ddl in [
        ('failed_login_count', 'INTEGER DEFAULT 0'),
        ('locked_until', 'TIMESTAMP'),
        ('last_failed_login', 'TIMESTAMP'),
    ]:
        try:
            conn.execute(f"ALTER TABLE users ADD COLUMN {col} {ddl}")
        except Exception:
            pass


def migrate_v27(conn):
    """Email-notification flags — a global kill switch (app_setting
    smtp_notifications_enabled) and a per-user opt-out
    (users.email_notifications_enabled). Both default to on so existing
    behavior is unchanged for existing installations."""
    try:
        conn.execute(
            "ALTER TABLE users ADD COLUMN email_notifications_enabled INTEGER DEFAULT 1"
        )
    except Exception:
        pass
    import json as json_mod
    existing = conn.execute(
        "SELECT key FROM app_settings WHERE key = 'smtp_notifications_enabled'"
    ).fetchone()
    if not existing:
        conn.execute(
            "INSERT INTO app_settings (key, value) VALUES ('smtp_notifications_enabled', ?)",
            (json_mod.dumps(True),)
        )


def migrate_v26(conn):
    """Add parent_id + author_user_id to work_order_notes for threaded
    conversation replies. Top-level notes have parent_id=NULL; every reply
    points at the root note it belongs to (flat one-level threading)."""
    try:
        conn.execute("ALTER TABLE work_order_notes ADD COLUMN parent_id INTEGER DEFAULT NULL")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE work_order_notes ADD COLUMN author_user_id INTEGER DEFAULT NULL")
    except Exception:
        pass
    conn.execute("CREATE INDEX IF NOT EXISTS idx_won_parent ON work_order_notes(parent_id)")
    # Backfill author_user_id where we can match by display_name or username
    conn.execute("""
        UPDATE work_order_notes
           SET author_user_id = (
               SELECT u.id FROM users u
                WHERE LOWER(u.display_name) = LOWER(work_order_notes.author)
                   OR LOWER(u.username) = LOWER(work_order_notes.author)
                LIMIT 1
           )
         WHERE author_user_id IS NULL AND author IS NOT NULL AND author != ''
    """)


def migrate_v25(conn):
    """Add is_sales_person flag to users. Sales people for work orders are now
    derived from users with this flag set (their username = email). Backfills
    from the legacy wo_salespeople app_setting by matching display_name or
    username so existing configurations keep working."""
    try:
        conn.execute("ALTER TABLE users ADD COLUMN is_sales_person INTEGER DEFAULT 0")
    except Exception:
        pass
    import json as json_mod
    row = conn.execute("SELECT value FROM app_settings WHERE key = 'wo_salespeople'").fetchone()
    if row:
        try:
            legacy = json_mod.loads(row['value'] or '[]') or []
        except Exception:
            legacy = []
        for sp in legacy:
            if not isinstance(sp, dict):
                continue
            name = str(sp.get('name', '')).strip()
            email = str(sp.get('email', '')).strip()
            if not name:
                continue
            # Match by display_name (case-insensitive), then username, then email
            matched = conn.execute(
                "UPDATE users SET is_sales_person = 1 "
                "WHERE LOWER(display_name) = LOWER(?) OR LOWER(username) = LOWER(?) "
                "OR (? != '' AND LOWER(username) = LOWER(?))",
                (name, name, email, email)
            )
            _ = matched  # rowcount not needed; silent backfill


# ┌──────────────────────────────────────────────┐
# │  MIGRATIONS REGISTRY — append new ones here  │
# └──────────────────────────────────────────────┘
MIGRATIONS = [
    (1, migrate_v1),
    (2, migrate_v2),
    (3, migrate_v3),
    (4, migrate_v4),
    (5, migrate_v5),
    (6, migrate_v6),
    (7, migrate_v7),
    (8, migrate_v8),
    (9, migrate_v9),
    (10, migrate_v10),
    (11, migrate_v11),
    (12, migrate_v12),
    (13, migrate_v13),
    (14, migrate_v14),
    (15, migrate_v15),
    (16, migrate_v16),
    (17, migrate_v17),
    (18, migrate_v18),
    (19, migrate_v19),
    (20, migrate_v20),
    (21, migrate_v21),
    (22, migrate_v22),
    (23, migrate_v23),
    (24, migrate_v24),
    (25, migrate_v25),
    (26, migrate_v26),
    (27, migrate_v27),
    (28, migrate_v28),
    (29, migrate_v29),
]


def init_db():
    """Run all pending migrations (with file lock to prevent Gunicorn worker races)."""
    import fcntl
    lock_path = os.path.join(DATA_DIR, '.migration_lock')
    lock_file = open(lock_path, 'w')
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX)

        conn = get_db()
        conn.execute('''
            CREATE TABLE IF NOT EXISTS schema_version (
                version INTEGER NOT NULL
            )
        ''')

        row = conn.execute("SELECT MAX(version) as v FROM schema_version").fetchone()
        current_version = row['v'] if row['v'] is not None else 0

        applied = 0
        for version, migrate_fn in MIGRATIONS:
            if version > current_version:
                print(f"  → Applying migration v{version}: {migrate_fn.__doc__.strip()}")
                migrate_fn(conn)
                conn.execute("INSERT INTO schema_version (version) VALUES (?)", (version,))
                conn.commit()
                applied += 1

        if applied:
            print(f"  ✓ {applied} migration(s) applied (now at v{MIGRATIONS[-1][0]})")
        else:
            print(f"  ✓ Database up to date (v{current_version})")

        conn.close()
    finally:
        fcntl.flock(lock_file, fcntl.LOCK_UN)
        lock_file.close()


# ══════════════════════════════════════════
#  SEARCH COLS
# ══════════════════════════════════════════

SEARCH_COLS = {
    'head': ['sku','location','head_engine','head_part','foundry_number','foundry','head_number','head_type','head_old_number','fitment_vehicle','sold_date','notes'],
    'engine': ['sku','location','engine_name','engine_head','engine_litre','engine_date_stamp','fitment_vehicle','sold_date','notes'],
    'transmission': ['sku','location','trans_date_code','trans_stamped_numbers','fitment_vehicle','sold_date','notes'],
}
ALL_SEARCH_COLS = list(set(col for cols in SEARCH_COLS.values() for col in cols))

ALL_FIELDS = [
    'category','sku','location','fitment_vehicle','notes','image_filename',
    'sold','sold_date','flagged','needs_audit','audit_note',
    'head_engine','head_part','foundry_number','foundry','head_number','head_type','head_old_number',
    'engine_name','engine_head','engine_litre','engine_date_stamp','engine_turns',
    'trans_gear_condition','trans_spins','trans_shifts','trans_date_code','trans_stamped_numbers',
]
INT_FIELDS = {'sold', 'flagged', 'needs_audit'}
RADIO_FIELDS = {'engine_turns', 'trans_spins', 'trans_shifts'}

# Importable fields per category (used by import system)
IMPORT_FIELDS = {
    'head': [
        ('sku', 'SKU'), ('head_old_number', 'Old SKU'), ('location', 'Location'), ('head_engine', 'Engine'),
        ('head_part', 'Part'), ('foundry_number', 'Foundry Number'), ('foundry', 'Foundry'),
        ('head_number', 'Head Number'), ('head_type', 'Type'),
        ('fitment_vehicle', 'Fitment Vehicle'), ('sold', 'Sold'), ('sold_date', 'Sold Date'),
        ('notes', 'Notes'),
    ],
    'engine': [
        ('sku', 'SKU'), ('location', 'Location'), ('engine_turns', 'Turns'),
        ('engine_name', 'Engine'), ('engine_head', 'Head'), ('engine_litre', 'Litre'),
        ('engine_date_stamp', 'Date Stamp'), ('fitment_vehicle', 'Fitment Vehicle'),
        ('sold', 'Sold'), ('sold_date', 'Sold Date'), ('notes', 'Notes'),
    ],
    'transmission': [
        ('sku', 'SKU'), ('location', 'Location'), ('trans_spins', 'Spins'),
        ('trans_shifts', 'Shifts'), ('trans_date_code', 'Date Code'),
        ('trans_stamped_numbers', 'Stamped Numbers'), ('fitment_vehicle', 'Fitment Vehicle'),
        ('sold', 'Sold'), ('sold_date', 'Sold Date'), ('notes', 'Notes'),
    ],
}

UPLOAD_TEMP = os.path.join(UPLOAD_DIR, 'temp')
os.makedirs(UPLOAD_TEMP, exist_ok=True)


# ══════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_image(file_field):
    if file_field and file_field.filename and allowed_file(file_field.filename):
        ext = file_field.filename.rsplit('.', 1)[1].lower()
        fname = f"{uuid.uuid4().hex}.{ext}"
        file_field.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
        return fname
    return ''


def save_image_resized(file_field, max_edge=2048, jpeg_quality=80):
    """Save an uploaded image re-encoded as JPEG with the long edge capped.
    Returns the stored filename, or '' if the file is not a valid image."""
    if not (file_field and file_field.filename and allowed_file(file_field.filename)):
        return ''
    from PIL import Image, ImageOps
    try:
        img = Image.open(file_field.stream)
        img = ImageOps.exif_transpose(img)  # respect phone-photo orientation
        if img.mode in ('RGBA', 'LA', 'P'):
            # Flatten transparency on white — JPEG has no alpha channel
            bg = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            bg.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
            img = bg
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        img.thumbnail((max_edge, max_edge), Image.LANCZOS)
        fname = f"{uuid.uuid4().hex}.jpg"
        img.save(os.path.join(app.config['UPLOAD_FOLDER'], fname),
                 format='JPEG', quality=jpeg_quality, optimize=True)
        return fname
    except Exception:
        return ''

def delete_image(filename):
    if filename:
        path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(path):
            os.remove(path)


def get_part_images(conn, part_id):
    """Get all images for a part, ordered by sort_order."""
    rows = conn.execute(
        "SELECT id, filename, sort_order FROM part_images WHERE part_id = ? ORDER BY sort_order, id",
        (part_id,)
    ).fetchall()
    return [dict(r) for r in rows]


def save_part_images(conn, part_id, files):
    """Save multiple uploaded images for a part."""
    max_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order), -1) FROM part_images WHERE part_id = ?", (part_id,)
    ).fetchone()[0]
    saved = []
    for f in files:
        fname = save_image(f)
        if fname:
            max_order += 1
            conn.execute(
                "INSERT INTO part_images (part_id, filename, sort_order) VALUES (?, ?, ?)",
                (part_id, fname, max_order)
            )
            saved.append(fname)
    return saved


def enrich_part_with_images(conn, part_dict):
    """Add images list and primary image_filename to a part dict."""
    images = get_part_images(conn, part_dict['id'])
    part_dict['images'] = images
    part_dict['image_filename'] = images[0]['filename'] if images else ''
    return part_dict


def assign_product_number(conn):
    """Get the next available product number and return it."""
    row = conn.execute("SELECT product_number FROM parts WHERE product_number LIKE 'WM%' AND product_number != '' ORDER BY CAST(REPLACE(REPLACE(product_number, 'WM-', ''), 'WM', '') AS INTEGER) DESC LIMIT 1").fetchone()
    if row and row['product_number']:
        try:
            num_str = row['product_number'].replace('WM-', '').replace('WM', '')
            next_num = int(num_str) + 1
        except ValueError:
            next_num = 1
    else:
        next_num = 1
    # Pad to 5 digits, but allow overflow past 99999 naturally
    if next_num < 100000:
        return f"WM{next_num:05d}"
    return f"WM{next_num}"

def form_val(key, default=''):
    val = request.form.get(key, default)
    if key in INT_FIELDS:
        try: return int(val)
        except (ValueError, TypeError): return 0
    return val


# ══════════════════════════════════════════
#  AUTH ROUTES
# ══════════════════════════════════════════

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect('/dashboard')

    # Pull turnstile config + branding once per request so GET and POST both see it
    conn_ts = get_db()
    ts_cfg = _get_setting(conn_ts, 'turnstile_config', {})
    branding_logo = _branding_filename(conn_ts)
    branding_width = _branding_logo_width(conn_ts)
    conn_ts.close()
    ts_ctx = {
        'turnstile_enabled': bool(ts_cfg.get('enabled')),
        'turnstile_site_key': ts_cfg.get('site_key', '') if ts_cfg.get('enabled') else '',
        'branding_logo_url': '/branding/logo' if branding_logo else '',
        'branding_logo_width': branding_width,
        'app_version': APP_VERSION,
    }

    if request.method == 'POST':
        username = ''
        password = ''
        turnstile_token = ''
        is_ajax = False
        try:
            data = request.get_json(force=True, silent=True)
            if data and isinstance(data, dict) and 'username' in data:
                username = data.get('username', '').strip()
                password = data.get('password', '')
                turnstile_token = data.get('cf-turnstile-response', '') or data.get('turnstile_token', '')
                is_ajax = True
        except Exception:
            pass

        if not is_ajax:
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            turnstile_token = request.form.get('cf-turnstile-response', '')
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        # Verify Turnstile first (no-op if not enabled)
        ok_ts, ts_err = _verify_turnstile(turnstile_token, request.remote_addr)
        if not ok_ts:
            if is_ajax:
                return jsonify({'error': ts_err or 'Challenge failed'}), 403
            return render_template('login.html', error=ts_err or 'Challenge failed', **ts_ctx)

        from datetime import datetime, timedelta
        conn = get_db()
        row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()

        # ── Check lockout state before verifying the password ──
        locked_msg = None
        if row and 'locked_until' in row.keys() and row['locked_until']:
            try:
                locked_at = datetime.fromisoformat(row['locked_until'].replace('T', ' ').split('.')[0])
            except Exception:
                locked_at = None
            if locked_at and locked_at > datetime.utcnow():
                remaining = int((locked_at - datetime.utcnow()).total_seconds() // 60) + 1
                locked_msg = f"Account locked — try again in {remaining} minute{'s' if remaining != 1 else ''} or contact an administrator."
            elif locked_at:
                # Lockout expired — clear it
                conn.execute(
                    "UPDATE users SET locked_until = NULL, failed_login_count = 0 WHERE id = ?",
                    (row['id'],)
                )
                conn.commit()

        # Always run a password hash, even if the user is missing or locked, so
        # the request takes roughly the same time regardless (timing-safe).
        _DUMMY_HASH = generate_password_hash('invalid', method='pbkdf2:sha256')
        if locked_msg:
            check_password_hash(_DUMMY_HASH, password)
            conn.close()
            if is_ajax:
                return jsonify({'error': locked_msg}), 423
            return render_template('login.html', error=locked_msg, **ts_ctx)

        if row and row['active'] and check_password_hash(row['password_hash'], password):
            # Success — reset counters
            conn.execute(
                "UPDATE users SET failed_login_count = 0, locked_until = NULL WHERE id = ?",
                (row['id'],)
            )
            conn.commit()
            conn.close()
            user = User(row['id'], row['username'], row['display_name'], row['role'], row['active'])
            login_user(user, remember=True)
            session.permanent = True
            # Always land on the dashboard after signing in — ignore any `next`
            # parameter so users start at a known screen regardless of where
            # the auth redirect came from.
            if is_ajax:
                return jsonify({'success': True, 'redirect': '/dashboard'})
            return redirect('/dashboard')

        # Failure — burn a hash for missing users so timing can't distinguish
        if not row:
            check_password_hash(_DUMMY_HASH, password)

        # Record the failure against an existing user (not unknown usernames
        # — we don't want to create lockouts for accounts that don't exist).
        attempts_left = None
        if row and row['active']:
            now_iso = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            new_count = (row['failed_login_count'] or 0) + 1 if 'failed_login_count' in row.keys() else 1
            lock_value = None
            if new_count >= LOGIN_FAIL_LIMIT:
                lock_value = (datetime.utcnow() + timedelta(minutes=LOCKOUT_MINUTES)).strftime('%Y-%m-%d %H:%M:%S')
            conn.execute(
                "UPDATE users SET failed_login_count = ?, last_failed_login = ?, locked_until = ? "
                "WHERE id = ?",
                (new_count, now_iso, lock_value, row['id'])
            )
            conn.commit()
            if lock_value:
                attempts_left = 0
            else:
                attempts_left = LOGIN_FAIL_LIMIT - new_count
        conn.close()

        if attempts_left == 0:
            err = f"Account locked for {LOCKOUT_MINUTES} minutes after {LOGIN_FAIL_LIMIT} failed attempts."
            if is_ajax:
                return jsonify({'error': err}), 423
            return render_template('login.html', error=err, **ts_ctx)

        err = 'Invalid username or password'
        if is_ajax:
            return jsonify({'error': err}), 401
        return render_template('login.html', error=err, **ts_ctx)

    return render_template('login.html', error=None, **ts_ctx)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/api/auth/me')
@login_required
def auth_me():
    return jsonify({
        'id': current_user.id,
        'username': current_user.username,
        'display_name': current_user.display_name,
        'role': current_user.role,
        'can_edit': current_user.can_edit,
        'can_view_audit': current_user.can_view_audit,
        'version': APP_VERSION,
    })


#
# Changelog — CHANGELOG.md is the single source of truth, shipped with the
# Docker image. The About tab fetches it at runtime so bumping the file
# automatically updates the in-app pane on the next build.
#
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
_CHANGELOG_PATH = os.path.join(_APP_DIR, 'CHANGELOG.md')
_CHANGELOG_CACHE = None


def _load_changelog():
    """Parse CHANGELOG.md into [{version, bullets: [...]} ...].
    Each `## vX.Y.Z` line starts a new version block; lines starting with `-`
    become bullet items. Bullet HTML is preserved as-authored so the About
    tab can render <strong>/<code>/<em> inline. Cached after first load."""
    global _CHANGELOG_CACHE
    if _CHANGELOG_CACHE is not None:
        return _CHANGELOG_CACHE
    if not os.path.exists(_CHANGELOG_PATH):
        _CHANGELOG_CACHE = []
        return _CHANGELOG_CACHE
    versions = []
    current = None
    header_re = re.compile(r'^##\s+v?([0-9][^\s]*)')
    try:
        with open(_CHANGELOG_PATH, encoding='utf-8') as fh:
            for raw in fh:
                line = raw.rstrip('\n')
                m = header_re.match(line)
                if m:
                    if current and current['bullets']:
                        versions.append(current)
                    current = {'version': m.group(1).strip(), 'bullets': []}
                    continue
                if current is not None:
                    stripped = line.lstrip()
                    if stripped.startswith('- '):
                        current['bullets'].append(stripped[2:].strip())
                    elif stripped.startswith('-'):
                        current['bullets'].append(stripped[1:].strip())
                    elif stripped and current['bullets']:
                        # Continuation of the previous bullet (indented line)
                        current['bullets'][-1] += ' ' + stripped
        if current and current['bullets']:
            versions.append(current)
    except Exception as e:
        app.logger.warning(f"Failed to parse CHANGELOG.md: {e}")
    _CHANGELOG_CACHE = versions
    return versions


@app.route('/api/changelog')
@login_required
def api_changelog():
    return jsonify({'version': APP_VERSION, 'entries': _load_changelog()})


@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def change_own_password():
    data = request.get_json()
    current_pw = data.get('current_password', '')
    new_pw = data.get('new_password', '')

    if not new_pw or len(new_pw) < PASSWORD_MIN_LENGTH:
        return jsonify({'error': f'New password must be at least {PASSWORD_MIN_LENGTH} characters'}), 400

    conn = get_db()
    row = conn.execute("SELECT password_hash FROM users WHERE id = ?", (current_user.id,)).fetchone()
    if not check_password_hash(row['password_hash'], current_pw):
        conn.close()
        return jsonify({'error': 'Current password is incorrect'}), 400

    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?",
                 (generate_password_hash(new_pw, method='pbkdf2:sha256'), current_user.id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ══════════════════════════════════════════
#  ADMIN — USER MANAGEMENT
# ══════════════════════════════════════════

_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def _valid_email(s):
    return bool(s) and bool(_EMAIL_RE.match(s))


@app.route('/api/users', methods=['GET'])
@admin_required
def list_users():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, username, display_name, role, active, is_sales_person, created_at "
        "FROM users ORDER BY created_at"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    data = request.get_json()
    username = data.get('username', '').strip()
    display_name = data.get('display_name', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'viewer')
    is_sales_person = 1 if bool(data.get('is_sales_person')) else 0

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    if not _valid_email(username):
        return jsonify({'error': 'Username must be a valid email address'}), 400
    if len(password) < PASSWORD_MIN_LENGTH:
        return jsonify({'error': f'Password must be at least {PASSWORD_MIN_LENGTH} characters'}), 400
    if role not in ('admin', 'editor', 'supervisor', 'viewer'):
        return jsonify({'error': 'Invalid role'}), 400

    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'Username already exists'}), 409

    pw_hash = generate_password_hash(password, method='pbkdf2:sha256')
    cur = conn.execute(
        "INSERT INTO users (username, display_name, password_hash, role, active, is_sales_person) "
        "VALUES (?, ?, ?, ?, 1, ?)",
        (username, display_name or username, pw_hash, role, is_sales_person)
    )
    conn.commit()
    row = conn.execute(
        "SELECT id, username, display_name, role, active, is_sales_person, created_at "
        "FROM users WHERE id = ?",
        (cur.lastrowid,)
    ).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


@app.route('/api/users/<int:uid>', methods=['PUT'])
@admin_required
def update_user(uid):
    data = request.get_json()
    conn = get_db()
    existing = conn.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'User not found'}), 404

    # Prevent demoting yourself or deactivating yourself
    if uid == current_user.id:
        if data.get('role') and data['role'] != 'admin':
            conn.close()
            return jsonify({'error': 'Cannot remove your own admin role'}), 400
        if 'active' in data and not data['active']:
            conn.close()
            return jsonify({'error': 'Cannot deactivate your own account'}), 400

    username = data.get('username', existing['username']).strip()
    display_name = data.get('display_name', existing['display_name']).strip()
    role = data.get('role', existing['role'])
    active = int(data.get('active', existing['active']))
    if 'is_sales_person' in data:
        is_sales_person = 1 if bool(data.get('is_sales_person')) else 0
    else:
        is_sales_person = existing['is_sales_person'] if 'is_sales_person' in existing.keys() else 0

    if role not in ('admin', 'editor', 'supervisor', 'viewer'):
        conn.close()
        return jsonify({'error': 'Invalid role'}), 400
    if not _valid_email(username):
        conn.close()
        return jsonify({'error': 'Username must be a valid email address'}), 400

    # Check username conflict
    conflict = conn.execute("SELECT id FROM users WHERE username = ? AND id != ?", (username, uid)).fetchone()
    if conflict:
        conn.close()
        return jsonify({'error': 'Username already taken'}), 409

    conn.execute(
        "UPDATE users SET username=?, display_name=?, role=?, active=?, is_sales_person=? WHERE id=?",
        (username, display_name, role, active, is_sales_person, uid)
    )

    # Optional password reset
    new_pw = data.get('password', '')
    if new_pw:
        if len(new_pw) < PASSWORD_MIN_LENGTH:
            conn.close()
            return jsonify({'error': f'Password must be at least {PASSWORD_MIN_LENGTH} characters'}), 400
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?",
                     (generate_password_hash(new_pw, method='pbkdf2:sha256'), uid))

    conn.commit()
    row = conn.execute(
        "SELECT id, username, display_name, role, active, is_sales_person, created_at "
        "FROM users WHERE id = ?",
        (uid,)
    ).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.route('/api/users/<int:uid>', methods=['DELETE'])
@admin_required
def delete_user(uid):
    if uid == current_user.id:
        return jsonify({'error': 'Cannot delete your own account'}), 400

    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE id = ?", (uid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'User not found'}), 404

    conn.execute("DELETE FROM users WHERE id = ?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/users/locked', methods=['GET'])
@admin_required
def list_locked_users():
    """Users whose lockout timestamp is still in the future. Drives the
    admin Locked Accounts widget on the dashboard."""
    from datetime import datetime
    now_iso = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    rows = conn.execute(
        "SELECT id, username, display_name, locked_until, failed_login_count, last_failed_login "
        "FROM users WHERE locked_until IS NOT NULL AND locked_until > ? "
        "ORDER BY locked_until DESC",
        (now_iso,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/users/<int:uid>/unlock', methods=['POST'])
@admin_required
def unlock_user(uid):
    """Admin unlock — clears the lockout timestamp and resets the failure
    counter so the user can sign in immediately."""
    conn = get_db()
    existing = conn.execute(
        "SELECT id, username FROM users WHERE id = ?", (uid,)
    ).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'User not found'}), 404
    conn.execute(
        "UPDATE users SET locked_until = NULL, failed_login_count = 0 WHERE id = ?",
        (uid,)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'username': existing['username']})


# ══════════════════════════════════════════
#  PAGE ROUTES
# ══════════════════════════════════════════

@app.route('/')
@app.route('/dashboard')
@app.route('/workorders')
@app.route('/workorders/archive')
@app.route('/parts')
@app.route('/parts/<slug>')
@login_required
def index(slug=None):
    return render_template('index.html')


_UPLOAD_NAME_RE = re.compile(r'^[a-f0-9]{32}\.(?:png|jpg|jpeg|gif|webp)$', re.IGNORECASE)


@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    # Defense-in-depth: send_from_directory already rejects traversal, but we
    # also require the filename to match the UUID-hex.<ext> pattern every save
    # path produces. Anything else (legacy, hand-crafted) → 404.
    if not _UPLOAD_NAME_RE.match(filename or ''):
        return '', 404
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# ══════════════════════════════════════════
#  PARTS API (all @login_required)
# ══════════════════════════════════════════

@app.route('/api/parts', methods=['GET'])
@login_required
def get_parts():
    category = request.args.get('category', '')
    search = request.args.get('search', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 24))

    conn = get_db()
    conds, params = [], []
    if category:
        conds.append("category = ?")
        params.append(category)
    if search:
        term = f"%{search}%"
        cols = ['sku', 'location', 'fitment_vehicle', 'notes', 'sold_date', 'custom_data']
        conds.append("(" + " OR ".join(f"{c} LIKE ?" for c in cols) + ")")
        params.extend([term] * len(cols))

    where = "WHERE " + " AND ".join(conds) if conds else ""

    # Sorting
    SORT_ALLOWED = {'sku', 'location', 'fitment_vehicle', 'updated_at', 'product_number',
                    'sold', 'flagged', 'needs_audit', 'posted_to_web'}
    sort_by = request.args.get('sort_by', 'sku')
    sort_dir = request.args.get('sort_dir', 'desc').lower()
    if sort_by not in SORT_ALLOWED:
        sort_by = 'sku'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'desc'
    reverse = sort_dir == 'desc'

    if sort_by in ('sku', 'location', 'fitment_vehicle', 'product_number'):
        rows = conn.execute(f"SELECT * FROM parts {where}", params).fetchall()
        parts = [dict(r) for r in rows]
        parts.sort(key=lambda p: _natural_sort_key(p.get(sort_by, '')), reverse=reverse)
        total = len(parts)
        offset = (page - 1) * per_page
        page_parts = parts[offset:offset + per_page]
    elif sort_by == 'posted_to_web':
        # Field lives inside the custom_data JSON blob (per-category toggle).
        # Truthy = any non-empty, non-'0' string.
        total = conn.execute(f"SELECT COUNT(*) as n FROM parts {where}", params).fetchone()['n']
        offset = (page - 1) * per_page
        rows = conn.execute(
            f"SELECT * FROM parts {where} "
            f"ORDER BY (CASE WHEN COALESCE(json_extract(custom_data, '$.posted_to_web'), '') "
            f"NOT IN ('', '0') THEN 1 ELSE 0 END) {sort_dir.upper()}, id DESC "
            f"LIMIT ? OFFSET ?",
            params + [per_page, offset]
        ).fetchall()
        page_parts = [dict(r) for r in rows]
    else:
        total = conn.execute(f"SELECT COUNT(*) as n FROM parts {where}", params).fetchone()['n']
        offset = (page - 1) * per_page
        rows = conn.execute(
            f"SELECT * FROM parts {where} ORDER BY {sort_by} {sort_dir.upper()} LIMIT ? OFFSET ?",
            params + [per_page, offset]
        ).fetchall()
        page_parts = [dict(r) for r in rows]

    # Batch-fetch first image for each part
    if page_parts:
        pids = [p['id'] for p in page_parts]
        placeholders = ','.join(['?'] * len(pids))
        img_rows = conn.execute(
            f"SELECT part_id, filename FROM part_images WHERE part_id IN ({placeholders}) ORDER BY sort_order, id",
            pids
        ).fetchall()
        # Build map: part_id -> first filename
        first_img = {}
        for r in img_rows:
            if r['part_id'] not in first_img:
                first_img[r['part_id']] = r['filename']
        for p in page_parts:
            p['image_filename'] = first_img.get(p['id'], '')

    conn.close()
    return jsonify({
        'parts': page_parts, 'total': total, 'page': page,
        'per_page': per_page, 'pages': max(1, (total + per_page - 1) // per_page)
    })


def _valid_category(conn, slug):
    """Check if a category slug exists (built-in or custom)."""
    return conn.execute("SELECT id FROM categories WHERE slug = ?", (slug,)).fetchone() is not None



@app.route('/api/parts', methods=['POST'])
@editor_required
def create_part():
    f = request.form
    category = f.get('category', '')

    conn = get_db()
    if not _valid_category(conn, category):
        conn.close()
        return jsonify({'error': 'Invalid category'}), 400

    import json as json_mod

    # Shared fields go in SQL columns; category-specific fields go in custom_data
    shared_keys = {'category', 'sku', 'location', 'fitment_vehicle', 'sold', 'sold_date', 'notes', 'flagged', 'needs_audit', 'audit_note'}
    fields = [fld for fld in ALL_FIELDS if fld != 'image_filename']
    vals = [form_val(fld) for fld in fields] + ['']

    # Build custom_data from category fields
    cat_fields = conn.execute(
        "SELECT field_key FROM category_fields WHERE category_slug = ? ORDER BY sort_order", (category,)
    ).fetchall()
    custom_data = {}
    for cf in cat_fields:
        k = cf['field_key']
        if k in shared_keys:
            continue  # these are already in SQL columns via form_val
        # Check both custom_ prefixed and direct form keys
        if f"custom_{k}" in f:
            custom_data[k] = f.get(f"custom_{k}", '')
        elif k in f:
            custom_data[k] = f.get(k, '')

    fields_with_cd = fields + ['image_filename', 'custom_data']
    vals_with_cd = vals[:-1] + ['', json_mod.dumps(custom_data)]
    placeholders = ','.join(['?'] * len(fields_with_cd))
    col_names = ','.join(fields_with_cd)
    cur = conn.execute(f"INSERT INTO parts ({col_names}) VALUES ({placeholders})", vals_with_cd)

    part_id = cur.lastrowid

    # Assign product number
    pn = assign_product_number(conn)
    conn.execute("UPDATE parts SET product_number = ? WHERE id = ?", (pn, part_id))

    images = request.files.getlist('images')
    if not images or not images[0].filename:
        images = request.files.getlist('image')
    save_part_images(conn, part_id, images)

    conn.commit()
    row = conn.execute("SELECT * FROM parts WHERE id = ?", (part_id,)).fetchone()
    part = enrich_part_with_images(conn, dict(row))
    conn.close()
    return jsonify(part), 201


@app.route('/api/parts/<int:pid>', methods=['GET'])
@login_required
def get_part(pid):
    conn = get_db()
    row = conn.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    part = enrich_part_with_images(conn, dict(row))
    conn.close()
    return jsonify(part)


@app.route('/api/parts/<int:pid>', methods=['PUT'])
@editor_required
def update_part(pid):
    conn = get_db()
    existing = conn.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    f = request.form
    category = f.get('category', existing['category'])
    if not _valid_category(conn, category):
        conn.close()
        return jsonify({'error': 'Invalid category'}), 400

    fields = [fld for fld in ALL_FIELDS if fld != 'image_filename']
    sets = ','.join(f"{fld}=?" for fld in fields)
    vals = []
    for fld in fields:
        vals.append(form_val(fld) if fld in request.form else existing[fld])

    # Handle custom_data - all categories use it now
    import json as json_mod
    shared_keys = {'category', 'sku', 'location', 'fitment_vehicle', 'sold', 'sold_date', 'notes', 'flagged', 'needs_audit', 'audit_note'}
    cat_fields = conn.execute(
        "SELECT field_key FROM category_fields WHERE category_slug = ? ORDER BY sort_order", (category,)
    ).fetchall()
    existing_cd = {}
    try:
        existing_cd = json_mod.loads(existing['custom_data'] or '{}')
    except Exception:
        pass
    for cf in cat_fields:
        k = cf['field_key']
        if k in shared_keys:
            continue
        if f"custom_{k}" in f:
            existing_cd[k] = f.get(f"custom_{k}", '')
        elif k in f:
            existing_cd[k] = f.get(k, '')
    sets += ",custom_data=?"
    vals.append(json_mod.dumps(existing_cd))

    sets += ",updated_at=CURRENT_TIMESTAMP"
    vals.append(pid)
    conn.execute(f"UPDATE parts SET {sets} WHERE id=?", vals)

    # Handle new image uploads (add to gallery)
    images = request.files.getlist('images')
    if not images or not images[0].filename:
        images = request.files.getlist('image')
    if images and images[0].filename:
        save_part_images(conn, pid, images)

    conn.commit()
    row = conn.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()
    part = enrich_part_with_images(conn, dict(row))
    conn.close()
    return jsonify(part)


@app.route('/api/parts/<int:pid>/images/<int:img_id>', methods=['DELETE'])
@editor_required
def delete_part_image(pid, img_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM part_images WHERE id = ? AND part_id = ?", (img_id, pid)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Image not found'}), 404
    delete_image(row['filename'])
    conn.execute("DELETE FROM part_images WHERE id = ?", (img_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/parts/<int:pid>', methods=['DELETE'])
@editor_required
def delete_part(pid):
    conn = get_db()
    existing = conn.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    # Delete all images
    imgs = conn.execute("SELECT filename FROM part_images WHERE part_id = ?", (pid,)).fetchall()
    for img in imgs:
        delete_image(img['filename'])
    conn.execute("DELETE FROM part_images WHERE part_id = ?", (pid,))
    delete_image(existing['image_filename'])  # legacy field
    conn.execute("DELETE FROM parts WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/parts/<int:pid>/audit', methods=['POST'])
@editor_required
def set_part_audit(pid):
    data = request.get_json() or {}
    needs = bool(data.get('needs_audit', False))
    note = str(data.get('audit_note', '') or '').strip()
    conn = get_db()
    row = conn.execute("SELECT id FROM parts WHERE id = ?", (pid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    conn.execute(
        "UPDATE parts SET needs_audit = ?, audit_note = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (1 if needs else 0, note if needs else '', pid)
    )
    conn.commit()
    conn.close()
    return jsonify({'needs_audit': 1 if needs else 0, 'audit_note': note if needs else ''})


@app.route('/api/parts/<int:pid>/flag', methods=['POST'])
@login_required
def toggle_flag(pid):
    conn = get_db()
    row = conn.execute("SELECT flagged FROM parts WHERE id = ?", (pid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    new_val = 0 if row['flagged'] else 1
    conn.execute("UPDATE parts SET flagged = ? WHERE id = ?", (new_val, pid))
    conn.commit()
    conn.close()
    return jsonify({'flagged': new_val})


@app.route('/api/stats', methods=['GET'])
@login_required
def get_stats():
    conn = get_db()
    cats = conn.execute("SELECT slug, name FROM categories ORDER BY sort_order").fetchall()
    stats = {}
    total = 0
    for cat in cats:
        n = conn.execute("SELECT COUNT(*) as n FROM parts WHERE category = ?", (cat['slug'],)).fetchone()['n']
        stats[cat['slug']] = n
        total += n
    stats['total'] = total
    conn.close()
    return jsonify(stats)


# ══════════════════════════════════════════
#  CATEGORIES API
# ══════════════════════════════════════════

@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    conn = get_db()
    cats = conn.execute("SELECT * FROM categories ORDER BY sort_order, id").fetchall()
    result = []
    for cat in cats:
        c = dict(cat)
        fields = conn.execute(
            "SELECT * FROM category_fields WHERE category_slug = ? ORDER BY sort_order, id",
            (c['slug'],)
        ).fetchall()
        c['fields'] = [dict(f) for f in fields]
        result.append(c)
    conn.close()
    return jsonify(result)


@app.route('/api/categories', methods=['POST'])
@admin_required
def create_category():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400

    # Generate slug
    slug = re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')
    if not slug:
        return jsonify({'error': 'Invalid name'}), 400

    color = data.get('color', '#4a8eff')
    icon = data.get('icon', '')
    fields_data = data.get('fields', [])

    conn = get_db()
    existing = conn.execute("SELECT id FROM categories WHERE slug = ?", (slug,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': 'A category with this name already exists'}), 409

    max_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) FROM categories").fetchone()[0]
    conn.execute(
        "INSERT INTO categories (slug, name, icon, color, sort_order, is_builtin) VALUES (?, ?, ?, ?, ?, 0)",
        (slug, name, icon, color, max_order + 1)
    )

    for i, f in enumerate(fields_data):
        conn.execute(
            "INSERT INTO category_fields (category_slug, field_key, field_label, field_type, radio_options, show_on_card, show_in_table, sort_order) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (slug, f.get('field_key', '').strip(), f.get('field_label', '').strip(),
             f.get('field_type', 'text'), f.get('radio_options', ''),
             int(f.get('show_on_card', 0)), int(f.get('show_in_table', 0)), i)
        )

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'slug': slug}), 201


@app.route('/api/categories/<slug>', methods=['PUT'])
@admin_required
def update_category(slug):
    conn = get_db()
    existing = conn.execute("SELECT * FROM categories WHERE slug = ?", (slug,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Category not found'}), 404

    data = request.get_json()
    name = data.get('name', existing['name']).strip()
    color = data.get('color', existing['color'])
    icon = data.get('icon', existing['icon'])

    conn.execute("UPDATE categories SET name=?, color=?, icon=? WHERE slug=?",
                 (name, color, icon, slug))

    # Replace fields if provided
    if 'fields' in data:
        conn.execute("DELETE FROM category_fields WHERE category_slug = ?", (slug,))
        for i, f in enumerate(data['fields']):
            conn.execute(
                "INSERT INTO category_fields (category_slug, field_key, field_label, field_type, radio_options, show_on_card, show_in_table, sort_order) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (slug, f.get('field_key', '').strip(), f.get('field_label', '').strip(),
                 f.get('field_type', 'text'), f.get('radio_options', ''),
                 int(f.get('show_on_card', 0)), int(f.get('show_in_table', 0)), i)
            )

    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/categories/<slug>', methods=['DELETE'])
@admin_required
def delete_category(slug):
    conn = get_db()
    existing = conn.execute("SELECT * FROM categories WHERE slug = ?", (slug,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Category not found'}), 404
    # Check for parts using this category
    count = conn.execute("SELECT COUNT(*) as n FROM parts WHERE category = ?", (slug,)).fetchone()['n']
    if count > 0:
        conn.close()
        return jsonify({'error': f'Cannot delete: {count} parts use this category. Remove or reassign them first.'}), 400

    conn.execute("DELETE FROM category_fields WHERE category_slug = ?", (slug,))
    conn.execute("DELETE FROM categories WHERE slug = ?", (slug,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ══════════════════════════════════════════
#  IMPORT API
# ══════════════════════════════════════════

# headChart fields found in Welsh export Description HTML.
# Order is fixed — used to build virtual column headers/rows.
HC_KEYS = ['Engine #', 'Head #', 'Block Part #', 'Litre', 'Vehicle', 'Date Stamp', 'Mileage']
HC_VIRTUAL_HEADERS = [f'[HC] {k}' for k in HC_KEYS] + ['[HC] Clean Description']


def _parse_head_chart(html):
    """Extract hC-Name/hC-Data pairs from a headChart div. Returns {key: value}."""
    if not html or 'headChart' not in html:
        return {}
    pairs = re.findall(
        r'hC-Name"[^>]*>\s*([^<]+?)\s*</div>\s*<div\s+class="hC-Data"[^>]*>\s*([^<]*?)\s*</div>',
        html, re.IGNORECASE | re.DOTALL,
    )
    out = {}
    for name, value in pairs:
        key = name.strip().rstrip(':').strip()
        out[key] = re.sub(r'\s+', ' ', value).strip()
    return out


def _strip_balanced_div(html, class_name):
    """Remove <div class="class_name">...</div> including nested divs."""
    pat = re.compile(r'<div\s+class="' + re.escape(class_name) + r'"[^>]*>', re.IGNORECASE)
    while True:
        m = pat.search(html)
        if not m:
            return html
        start = m.start()
        depth = 1
        i = m.end()
        tag_re = re.compile(r'<(/?)div\b[^>]*>', re.IGNORECASE)
        while depth > 0:
            tm = tag_re.search(html, i)
            if not tm:
                return html[:start]  # unbalanced — drop tail
            depth += -1 if tm.group(1) else 1
            i = tm.end()
        html = html[:start] + html[i:]


def _clean_description(html):
    """Strip headChart div, makeoffer block, images, and HTML tags from a description."""
    if not html:
        return ''
    s = _strip_balanced_div(html, 'headChart')
    s = _strip_balanced_div(s, 'makeoffer')
    s = re.sub(r'<img\b[^>]*>', '', s, flags=re.IGNORECASE)
    s = re.sub(r'<br\s*/?>', '\n', s, flags=re.IGNORECASE)
    s = re.sub(r'</p>\s*<p[^>]*>', '\n\n', s, flags=re.IGNORECASE)
    s = re.sub(r'<[^>]+>', '', s)
    s = re.sub(r'[ \t]+', ' ', s)
    s = re.sub(r'\n{3,}', '\n\n', s)
    return s.strip()


def _detect_head_chart_col(all_rows, header_row=0):
    """Find index of the column whose data cells contain 'headChart'. Returns -1 if none."""
    if len(all_rows) <= header_row + 1:
        return -1
    width = len(all_rows[header_row])
    for col in range(width):
        for r in all_rows[header_row + 1: header_row + 200]:
            if col < len(r) and 'headChart' in (r[col] or ''):
                return col
    return -1


def _apply_head_chart(all_rows, hc_col, header_row=0):
    """Append virtual HC columns to every row. Mutates and returns all_rows."""
    if hc_col is None or hc_col < 0 or not all_rows:
        return all_rows
    for i, row in enumerate(all_rows):
        if i == header_row:
            row.extend(HC_VIRTUAL_HEADERS)
        elif i < header_row:
            row.extend([''] * len(HC_VIRTUAL_HEADERS))
        else:
            html = row[hc_col] if hc_col < len(row) else ''
            parsed = _parse_head_chart(html)
            row.extend([parsed.get(k, '') for k in HC_KEYS])
            row.append(_clean_description(html))
    return all_rows


def _read_file_rows(temp_path, sheet_name=None, hc_col=None, header_row=0):
    """Read all rows from a CSV or Excel file. Returns list of string lists.
    If hc_col is a non-negative int, append virtual HC columns parsed from that column."""
    import csv as csv_module
    ext = temp_path.rsplit('.', 1)[-1].lower()
    all_rows = []
    if ext == 'csv':
        with open(temp_path, 'r', encoding='utf-8-sig', errors='replace') as f:
            for row in csv_module.reader(f):
                all_rows.append([str(c).strip() for c in row])
    else:
        from openpyxl import load_workbook
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c) if c is not None else '' for c in row])
        wb.close()
    if hc_col is not None and hc_col >= 0:
        _apply_head_chart(all_rows, hc_col, header_row)
    return all_rows

@app.route('/api/import/fields', methods=['GET'])
@login_required
def import_fields():
    """Return importable field definitions per category."""
    return jsonify({cat: [{'key': k, 'label': l} for k, l in fields]
                    for cat, fields in IMPORT_FIELDS.items()})


@app.route('/api/import/upload', methods=['POST'])
@admin_required
def import_upload():
    """Upload an Excel or CSV file, return sheet names, column headers, and sample rows."""
    import csv as csv_module
    from io import StringIO

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xls', 'xlsm', 'csv'):
        return jsonify({'error': 'File must be .xlsx, .xls, .xlsm, or .csv'}), 400

    # Save temp file
    temp_name = f"{uuid.uuid4().hex}.{ext}"
    temp_path = os.path.join(UPLOAD_TEMP, temp_name)
    file.save(temp_path)

    try:
        result = {'temp_file': temp_name, 'sheets': []}

        def _detect_and_pack(name, rows_data):
            hc_col = _detect_head_chart_col(rows_data, header_row=0)
            sheet = {
                'name': name,
                'headers': rows_data[0],
                'sample_rows': rows_data[1:],
                'col_count': len(rows_data[0]),
                'headchart_col': hc_col,
                'headchart_keys': HC_VIRTUAL_HEADERS if hc_col >= 0 else [],
            }
            if hc_col >= 0:
                _apply_head_chart(rows_data, hc_col, 0)
                sheet['headers'] = rows_data[0]
                sheet['sample_rows'] = rows_data[1:]
                sheet['col_count'] = len(rows_data[0])
            return sheet

        if ext == 'csv':
            with open(temp_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv_module.reader(f)
                rows_data = []
                for row in reader:
                    rows_data.append([str(c).strip() for c in row])
                    if len(rows_data) > 25:
                        break
            if rows_data:
                result['sheets'].append(_detect_and_pack('CSV', rows_data))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(temp_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data = []
                for row in ws.iter_rows(values_only=True):
                    rows_data.append([str(c) if c is not None else '' for c in row])
                    if len(rows_data) > 25:
                        break
                if not rows_data:
                    continue
                result['sheets'].append(_detect_and_pack(sheet_name, rows_data))
            wb.close()

        return jsonify(result)
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({'error': f'Failed to read file: {str(e)}'}), 400


@app.route('/api/import/preview', methods=['POST'])
@admin_required
def import_preview():
    """Dry run: apply column mapping and return preview of what would be imported."""
    data = request.get_json()
    temp_file = data.get('temp_file', '')
    sheet_name = data.get('sheet', '')
    category = data.get('category', '')
    mapping = data.get('mapping', {})
    header_row = data.get('header_row', 0)
    hc_col = data.get('headchart_col', -1)
    try:
        hc_col = int(hc_col)
    except (TypeError, ValueError):
        hc_col = -1

    # UI prefixes non-shared field keys with 'custom_' for form collision avoidance;
    # IMPORT_FIELDS uses bare keys, so normalize here.
    mapping = {(k[7:] if k.startswith('custom_') else k): v for k, v in mapping.items()}

    if category not in IMPORT_FIELDS:
        return jsonify({'error': 'Invalid category'}), 400

    temp_path = os.path.join(UPLOAD_TEMP, temp_file)
    if not os.path.exists(temp_path):
        return jsonify({'error': 'Uploaded file not found. Please re-upload.'}), 404

    try:
        all_rows = _read_file_rows(temp_path, sheet_name, hc_col=hc_col, header_row=header_row)
    except Exception as e:
        return jsonify({'error': f'Failed to read file: {str(e)}'}), 400

    # Data rows start after header
    data_rows = all_rows[header_row + 1:]

    # Build preview
    field_labels = {k: l for k, l in IMPORT_FIELDS[category]}
    preview_rows = []
    errors = []

    for row_idx, row in enumerate(data_rows):
        # Skip completely empty rows
        if all(c.strip() == '' for c in row):
            continue

        mapped = {}
        for db_field, col_idx_str in mapping.items():
            col_idx = int(col_idx_str)
            if col_idx < 0:
                continue
            value = row[col_idx] if col_idx < len(row) else ''
            # Handle toggle fields (sold)
            if db_field in INT_FIELDS:
                value_lower = value.strip().lower()
                if value_lower in ('yes', 'y', '1', 'true'):
                    mapped[db_field] = 1
                elif value_lower in ('no', 'n', '0', 'false', ''):
                    mapped[db_field] = 0
                else:
                    mapped[db_field] = 0
                    errors.append(f"Row {row_idx + 1}: '{db_field}' value '{value}' not recognized, defaulting to No")
            elif db_field in RADIO_FIELDS:
                mapped[db_field] = value.strip().lower() if value.strip() else 'untested'
            else:
                mapped[db_field] = value.strip()

        # Check if row has any meaningful data
        text_vals = [v for k, v in mapped.items() if k not in INT_FIELDS and k not in RADIO_FIELDS and v]
        if not text_vals:
            continue

        mapped['_row_num'] = row_idx + header_row + 2  # Excel row number (1-based + header)
        preview_rows.append(mapped)

    return jsonify({
        'category': category,
        'field_labels': field_labels,
        'fields': [k for k, l in IMPORT_FIELDS[category]],
        'preview': preview_rows,
        'total': len(preview_rows),
        'warnings': errors,
    })


@app.route('/api/import/execute', methods=['POST'])
@admin_required
def import_execute():
    """Execute the import: insert all mapped rows into the database."""
    data = request.get_json()
    temp_file = data.get('temp_file', '')
    sheet_name = data.get('sheet', '')
    category = data.get('category', '')
    mapping = data.get('mapping', {})
    header_row = data.get('header_row', 0)
    hc_col = data.get('headchart_col', -1)
    try:
        hc_col = int(hc_col)
    except (TypeError, ValueError):
        hc_col = -1

    mapping = {(k[7:] if k.startswith('custom_') else k): v for k, v in mapping.items()}

    if category not in IMPORT_FIELDS:
        return jsonify({'error': 'Invalid category'}), 400

    temp_path = os.path.join(UPLOAD_TEMP, temp_file)
    if not os.path.exists(temp_path):
        return jsonify({'error': 'Uploaded file not found. Please re-upload.'}), 404

    try:
        all_rows = _read_file_rows(temp_path, sheet_name, hc_col=hc_col, header_row=header_row)
    except Exception as e:
        return jsonify({'error': f'Failed to read file: {str(e)}'}), 400

    data_rows = all_rows[header_row + 1:]

    # Determine all db columns to insert
    valid_fields = {k for k, l in IMPORT_FIELDS[category]}
    mapped_fields = [f for f in mapping.keys() if f in valid_fields and int(mapping[f]) >= 0]

    conn = get_db()
    inserted = 0
    errors = []

    for row_idx, row in enumerate(data_rows):
        if all((c.strip() == '' if isinstance(c, str) else c == '') for c in row):
            continue

        vals = {'category': category}
        has_data = False

        for db_field in mapped_fields:
            col_idx = int(mapping[db_field])
            value = row[col_idx] if col_idx < len(row) else ''
            value = value.strip() if isinstance(value, str) else value

            if db_field in INT_FIELDS:
                value_lower = str(value).strip().lower()
                vals[db_field] = 1 if value_lower in ('yes', 'y', '1', 'true') else 0
            elif db_field in RADIO_FIELDS:
                vals[db_field] = str(value).strip().lower() if str(value).strip() else 'untested'
            else:
                vals[db_field] = value
                if value:
                    has_data = True

        if not has_data:
            continue

        # Build insert
        all_cols = ['category'] + mapped_fields
        placeholders = ','.join(['?'] * len(all_cols))
        col_names = ','.join(all_cols)
        col_vals = [vals.get(c, '') for c in all_cols]

        try:
            cur = conn.execute(f"INSERT INTO parts ({col_names}) VALUES ({placeholders})", col_vals)
            pn = assign_product_number(conn)
            conn.execute("UPDATE parts SET product_number = ? WHERE id = ?", (pn, cur.lastrowid))
            inserted += 1
        except Exception as e:
            errors.append(f"Row {row_idx + header_row + 2}: {str(e)}")

    conn.commit()
    conn.close()

    # Cleanup temp file
    try:
        os.remove(temp_path)
    except OSError:
        pass

    return jsonify({
        'success': True,
        'inserted': inserted,
        'errors': errors,
    })


# ══════════════════════════════════════════
#  EXPORT API
# ══════════════════════════════════════════

@app.route('/api/export/csv', methods=['GET'])
@admin_required
def export_csv():
    """Export entire inventory as CSV."""
    import csv as csv_module
    from io import StringIO

    conn = get_db()
    rows = conn.execute("SELECT * FROM parts ORDER BY category, sku").fetchall()

    # Get all categories and their fields for headers
    all_cats = conn.execute("SELECT slug FROM categories ORDER BY sort_order").fetchall()
    all_cat_fields = {}
    all_custom_keys = []  # ordered unique custom field keys
    for cat in all_cats:
        cfs = conn.execute(
            "SELECT field_key, field_label FROM category_fields WHERE category_slug = ? ORDER BY sort_order",
            (cat['slug'],)
        ).fetchall()
        all_cat_fields[cat['slug']] = cfs
        for cf in cfs:
            if cf['field_key'] not in ('sku', 'location', 'fitment_vehicle', 'sold', 'sold_date', 'notes') and cf['field_key'] not in [k for k, _ in all_custom_keys]:
                all_custom_keys.append((cf['field_key'], cf['field_label']))
    conn.close()

    import json as json_mod
    shared_cols = [
        ('product_number', 'Product Number'),
        ('category', 'Category'),
        ('sku', 'SKU'),
        ('location', 'Location'),
        ('fitment_vehicle', 'Fitment Vehicle'),
        ('flagged', 'Flagged'),
        ('sold', 'Sold'),
        ('sold_date', 'Sold Date'),
        ('notes', 'Notes'),
    ]
    all_cols = shared_cols + all_custom_keys

    output = StringIO()
    writer = csv_module.writer(output)
    writer.writerow([label for _, label in all_cols])

    for row in rows:
        r = dict(row)
        cd = {}
        try:
            cd = json_mod.loads(r.get('custom_data', '{}') or '{}')
        except Exception:
            pass
        csv_row = []
        for key, _ in all_cols:
            if key in ('product_number', 'category', 'sku', 'location', 'fitment_vehicle', 'sold_date', 'notes'):
                csv_row.append(r.get(key, '') or '')
            elif key == 'sold':
                csv_row.append('Yes' if r.get('sold') else 'No')
            elif key == 'flagged':
                csv_row.append('Yes' if r.get('flagged') else 'No')
            else:
                # Custom data field — check custom_data first, then legacy SQL column
                val = cd.get(key, '') or r.get(key, '') or ''
                csv_row.append(str(val))
        writer.writerow(csv_row)

    csv_data = output.getvalue()
    output.close()

    return Response(
        csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=warehouse-export.csv'}
    )


# ══════════════════════════════════════════
#  LABEL & QR CODE API
# ══════════════════════════════════════════

def _generate_qr_png(data_str):
    """Generate a QR code as PNG bytes."""
    import qrcode
    from io import BytesIO
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=1)
    qr.add_data(data_str)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.getvalue()


def _get_label_fields(conn, part):
    """Get printable fields for a part (excluding sold, sold_date, notes, images)."""
    import json as json_mod
    cat = part['category']
    exclude_keys = {'sold', 'sold_date', 'notes'}

    fields = []
    fields.append(('Product #', part.get('product_number', '')))
    fields.append(('Category', getCatName(conn, cat)))

    cat_field_rows = conn.execute(
        "SELECT field_key, field_label, field_type FROM category_fields WHERE category_slug = ? ORDER BY sort_order",
        (cat,)
    ).fetchall()

    cd = {}
    try:
        cd = json_mod.loads(part.get('custom_data', '{}') or '{}')
    except Exception:
        pass

    for cf in cat_field_rows:
        k, label, ftype = cf['field_key'], cf['field_label'], cf['field_type']
        if k in exclude_keys:
            continue
        # Get value from custom_data first, then SQL column fallback
        val = cd.get(k, '') or part.get(k, '') or ''
        if ftype == 'toggle':
            val = 'Yes' if val and val not in ('0', 'No', 'false') else ''
        elif ftype == 'radio' and val:
            val = val.capitalize()
        if val:
            fields.append((label, str(val)))

    return fields


def getCatName(conn, slug):
    row = conn.execute("SELECT name FROM categories WHERE slug = ?", (slug,)).fetchone()
    return row['name'] if row else slug


@app.route('/api/parts/<int:pid>/qr.png')
@login_required
def part_qr_code(pid):
    conn = get_db()
    row = conn.execute("SELECT product_number FROM parts WHERE id = ?", (pid,)).fetchone()
    conn.close()
    if not row:
        return 'Not found', 404
    png = _generate_qr_png(row['product_number'])
    return Response(png, mimetype='image/png')


@app.route('/api/parts/<int:pid>/label.pdf')
@login_required
def part_label_pdf(pid):
    from reportlab.lib.pagesizes import inch
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.utils import ImageReader
    from io import BytesIO

    conn = get_db()
    row = conn.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    part = dict(row)
    label_fields = _get_label_fields(conn, part)
    conn.close()

    # 4" x 1" label
    w, h = 4 * inch, 1 * inch
    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=(w, h))

    # QR code on the left
    qr_png = _generate_qr_png(part.get('product_number', ''))
    qr_img = ImageReader(BytesIO(qr_png))
    qr_size = 0.82 * inch
    c.drawImage(qr_img, 4, (h - qr_size) / 2, qr_size, qr_size)

    # Text fields on the right
    x_start = qr_size + 10
    text_area_w = w - x_start - 6
    y = h - 13

    # Product number bold
    c.setFont("Helvetica-Bold", 9)
    pn = part.get('product_number', '')
    c.drawString(x_start, y, pn)
    y -= 10

    # SKU bold same size
    sku = part.get('sku', '')
    if sku:
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x_start, y, sku)
        y -= 10

    # Remaining fields in two columns
    remaining = [(lbl, val) for lbl, val in label_fields if lbl not in ('Product #', 'SKU')]
    col_width = text_area_w / 2 - 2
    font_size = 7
    line_h = 7
    c.setFont("Helvetica", font_size)

    col1_x = x_start
    col2_x = x_start + text_area_w / 2
    mid = (len(remaining) + 1) // 2
    col1_items = remaining[:mid]
    col2_items = remaining[mid:]

    col_y = y
    for item in col1_items:
        text = f"{item[0]}: {item[1]}"
        while c.stringWidth(text, "Helvetica", font_size) > col_width and len(text) > 8:
            text = text[:-4] + '…'
        c.drawString(col1_x, col_y, text)
        col_y -= line_h
        if col_y < 3:
            break

    col_y = y
    for item in col2_items:
        text = f"{item[0]}: {item[1]}"
        while c.stringWidth(text, "Helvetica", font_size) > col_width and len(text) > 8:
            text = text[:-4] + '…'
        c.drawString(col2_x, col_y, text)
        col_y -= line_h
        if col_y < 3:
            break

    c.save()
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'inline; filename=label-{pn}.pdf'}
    )


@app.route('/api/labels/batch', methods=['POST'])
@login_required
def batch_labels_pdf():
    """Generate a multi-label PDF for multiple parts."""
    from reportlab.lib.pagesizes import inch, letter
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.utils import ImageReader
    from io import BytesIO

    data = request.get_json()
    part_ids = data.get('ids', [])
    if not part_ids:
        return jsonify({'error': 'No parts specified'}), 400

    conn = get_db()

    # Label dimensions
    lw, lh = 4 * inch, 1 * inch
    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=(lw, lh))

    for i, pid in enumerate(part_ids):
        row = conn.execute("SELECT * FROM parts WHERE id = ?", (pid,)).fetchone()
        if not row:
            continue
        part = dict(row)
        label_fields = _get_label_fields(conn, part)

        if i > 0:
            c.showPage()

        # QR code
        qr_png = _generate_qr_png(part.get('product_number', ''))
        qr_img = ImageReader(BytesIO(qr_png))
        qr_size = 0.82 * inch
        c.drawImage(qr_img, 4, (lh - qr_size) / 2, qr_size, qr_size)

        # Text
        x_start = qr_size + 10
        text_area_w = lw - x_start - 6
        y = lh - 13

        c.setFont("Helvetica-Bold", 9)
        pn = part.get('product_number', '')
        c.drawString(x_start, y, pn)
        y -= 10

        sku = part.get('sku', '')
        if sku:
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x_start, y, sku)
            y -= 10

        # Remaining fields in two columns
        remaining = [(lbl, val) for lbl, val in label_fields if lbl not in ('Product #', 'SKU')]
        col_width = text_area_w / 2 - 2
        font_size = 7
        line_h = 7
        c.setFont("Helvetica", font_size)

        col1_x = x_start
        col2_x = x_start + text_area_w / 2
        mid = (len(remaining) + 1) // 2
        col1_items = remaining[:mid]
        col2_items = remaining[mid:]

        col_y = y
        for item in col1_items:
            text = f"{item[0]}: {item[1]}"
            while c.stringWidth(text, "Helvetica", font_size) > col_width and len(text) > 8:
                text = text[:-4] + '…'
            c.drawString(col1_x, col_y, text)
            col_y -= line_h
            if col_y < 3:
                break

        col_y = y
        for item in col2_items:
            text = f"{item[0]}: {item[1]}"
            while c.stringWidth(text, "Helvetica", font_size) > col_width and len(text) > 8:
                text = text[:-4] + '…'
            c.drawString(col2_x, col_y, text)
            col_y -= line_h
            if col_y < 3:
                break

    c.save()
    conn.close()
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'inline; filename=labels-batch.pdf'}
    )


# ══════════════════════════════════════════
#  APP SETTINGS (key/value, JSON-encoded values)
# ══════════════════════════════════════════

SETTINGS_KEYS = {'wo_locations', 'wo_salespeople', 'wo_priorities', 'smtp_config', 'turnstile_config', 'branding', 'smtp_notifications_enabled'}


def _get_setting(conn, key, default):
    import json as json_mod
    row = conn.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    if not row:
        return default
    try:
        return json_mod.loads(row['value'])
    except Exception:
        return default


def _set_setting(conn, key, value):
    import json as json_mod
    existing = conn.execute("SELECT key FROM app_settings WHERE key = ?", (key,)).fetchone()
    v = json_mod.dumps(value)
    if existing:
        conn.execute("UPDATE app_settings SET value = ? WHERE key = ?", (v, key))
    else:
        conn.execute("INSERT INTO app_settings (key, value) VALUES (?, ?)", (key, v))


_PRIORITY_DEFAULT_COLORS = {
    'Next Day Air': '#fff3e5',
}
# Swap old defaults forward so stored data picks up a new default automatically
_PRIORITY_DEPRECATED_COLORS = {
    'Next Day Air': {'#fef9c3'},
}


def _normalize_priority(p):
    """Accept either a legacy string or a {name,color} dict; return normalized dict."""
    if isinstance(p, dict):
        name = str(p.get('name', '')).strip()
        color = str(p.get('color', '') or '').strip()
    else:
        name = str(p).strip()
        color = _PRIORITY_DEFAULT_COLORS.get(name, '')
    # Only accept 3/4/6/8-digit hex colors; empty string disables coloring
    if color and not re.match(r'^#[0-9A-Fa-f]{3,8}$', color):
        color = ''
    # Rewrite superseded defaults to the new one so the value in the DB migrates forward
    if name in _PRIORITY_DEPRECATED_COLORS and color.lower() in _PRIORITY_DEPRECATED_COLORS[name]:
        color = _PRIORITY_DEFAULT_COLORS.get(name, '')
    return {'name': name, 'color': color}


def _derive_salespeople(conn):
    """Salespeople for work orders are users flagged is_sales_person.
    Returns [{name: display_name or username, email: username, user_id}, ...]."""
    rows = conn.execute(
        "SELECT id, username, display_name FROM users "
        "WHERE is_sales_person = 1 AND active = 1 "
        "ORDER BY COALESCE(NULLIF(display_name, ''), username) COLLATE NOCASE"
    ).fetchall()
    out = []
    for r in rows:
        display = (r['display_name'] or '').strip() or r['username']
        out.append({'name': display, 'email': r['username'], 'user_id': r['id']})
    return out


@app.route('/api/settings/work-order', methods=['GET'])
@login_required
def get_wo_settings():
    """Return work-order related settings (locations, salespeople, priorities).
    Salespeople are derived from users with is_sales_person=1; emails are stripped
    for non-admins."""
    conn = get_db()
    locations = _get_setting(conn, 'wo_locations', [])
    salespeople = _derive_salespeople(conn)
    raw_priorities = _get_setting(conn, 'wo_priorities', ['Normal', 'Next Day Air'])
    priorities = [_normalize_priority(p) for p in raw_priorities if _normalize_priority(p)['name']]
    conn.close()
    if not current_user.is_admin:
        salespeople = [{'name': s.get('name', '')} for s in salespeople]
    return jsonify({
        'locations': locations,
        'salespeople': salespeople,
        'priorities': priorities,
    })


@app.route('/api/settings/work-order', methods=['PUT'])
@admin_required
def update_wo_settings():
    data = request.get_json() or {}
    conn = get_db()
    if 'locations' in data:
        locs = [str(x).strip() for x in (data['locations'] or []) if str(x).strip()]
        _set_setting(conn, 'wo_locations', locs)
    # 'salespeople' is now derived from the users table (is_sales_person flag).
    # Silently ignore any client-supplied list so stale callers don't error.
    if 'priorities' in data:
        pr = []
        for x in (data['priorities'] or []):
            n = _normalize_priority(x)
            if n['name']:
                pr.append(n)
        _set_setting(conn, 'wo_priorities', pr)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/settings/smtp', methods=['GET'])
@admin_required
def get_smtp_settings():
    conn = get_db()
    cfg = _get_setting(conn, 'smtp_config', {})
    conn.close()
    # Mask password in response
    masked = dict(cfg)
    if masked.get('password'):
        masked['password'] = '********'
    return jsonify(masked)


@app.route('/api/settings/smtp', methods=['PUT'])
@admin_required
def update_smtp_settings():
    data = request.get_json() or {}
    conn = get_db()
    existing = _get_setting(conn, 'smtp_config', {})
    cfg = {
        'host': str(data.get('host', existing.get('host', ''))).strip(),
        'port': int(data.get('port', existing.get('port', 587)) or 587),
        'username': str(data.get('username', existing.get('username', ''))).strip(),
        'use_tls': bool(data.get('use_tls', existing.get('use_tls', True))),
        'from_email': str(data.get('from_email', existing.get('from_email', ''))).strip(),
        'from_name': str(data.get('from_name', existing.get('from_name', 'Warehouse Manager'))).strip(),
    }
    # Only update password if a non-empty non-masked value is provided
    pw = data.get('password', None)
    if pw is None or pw == '********':
        cfg['password'] = existing.get('password', '')
    else:
        cfg['password'] = str(pw)
    _set_setting(conn, 'smtp_config', cfg)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ── Cloudflare Turnstile ──

def _verify_turnstile(token, remote_ip=None):
    """Verify a Turnstile response token. Returns (ok: bool, error: str|None).
    If Turnstile is not enabled in settings, returns (True, None)."""
    import urllib.request
    import urllib.parse
    import json as json_mod

    conn = get_db()
    cfg = _get_setting(conn, 'turnstile_config', {})
    conn.close()

    if not cfg.get('enabled'):
        return True, None
    secret = (cfg.get('secret_key') or '').strip()
    if not secret:
        return False, 'Turnstile enabled but secret not configured'
    if not token:
        return False, 'Please complete the challenge'

    try:
        data = urllib.parse.urlencode({
            'secret': secret,
            'response': token,
            'remoteip': remote_ip or '',
        }).encode()
        req = urllib.request.Request(
            'https://challenges.cloudflare.com/turnstile/v0/siteverify',
            data=data, method='POST'
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json_mod.loads(resp.read().decode())
        if result.get('success'):
            return True, None
        errs = ','.join(result.get('error-codes') or []) or 'verification failed'
        return False, f"Challenge verification failed: {errs}"
    except Exception as e:
        return False, f"Turnstile verification error: {e}"


@app.route('/api/auth/turnstile-config', methods=['GET'])
def public_turnstile_config():
    """Public endpoint — returns site_key + enabled flag for the login page."""
    conn = get_db()
    cfg = _get_setting(conn, 'turnstile_config', {})
    conn.close()
    enabled = bool(cfg.get('enabled'))
    return jsonify({
        'enabled': enabled,
        'site_key': cfg.get('site_key', '') if enabled else '',
    })


@app.route('/api/settings/turnstile', methods=['GET'])
@admin_required
def get_turnstile_settings():
    conn = get_db()
    cfg = _get_setting(conn, 'turnstile_config', {'enabled': False, 'site_key': '', 'secret_key': ''})
    conn.close()
    masked = {
        'enabled': bool(cfg.get('enabled', False)),
        'site_key': cfg.get('site_key', ''),
        'secret_key': '********' if cfg.get('secret_key') else '',
    }
    return jsonify(masked)


@app.route('/api/settings/turnstile', methods=['PUT'])
@admin_required
def update_turnstile_settings():
    data = request.get_json() or {}
    conn = get_db()
    existing = _get_setting(conn, 'turnstile_config', {})
    cfg = {
        'enabled': bool(data.get('enabled', existing.get('enabled', False))),
        'site_key': str(data.get('site_key', existing.get('site_key', ''))).strip(),
    }
    sk = data.get('secret_key', None)
    if sk is None or sk == '********':
        cfg['secret_key'] = existing.get('secret_key', '')
    else:
        cfg['secret_key'] = str(sk).strip()
    _set_setting(conn, 'turnstile_config', cfg)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


#
# Custom branding (login-screen logo)
# ─────────────────────────────────────────────────────────────
# Stored as app_setting branding = {'logo_filename': '<uuid>.<ext>'}.
# Files live in UPLOAD_DIR/branding/ so they don't collide with part uploads.
#

BRANDING_DIR = os.path.join(UPLOAD_DIR, 'branding')
os.makedirs(BRANDING_DIR, exist_ok=True)
# Branding uploads are admin-only (decorator-enforced). PNG is validated by
# Pillow; SVG is sanitized to strip <script>, on* handlers, external resource
# refs, and dangerous URL schemes. The /branding/logo response also carries a
# strict CSP so any smuggled payload would be neutered in the browser.
BRANDING_EXTS = {'png', 'svg'}
_BRANDING_NAME_RE = re.compile(r'^logo-[a-f0-9]{32}\.(?:png|svg)$', re.IGNORECASE)


def _branding_filename(conn):
    cfg = _get_setting(conn, 'branding', {})
    return str((cfg or {}).get('logo_filename', '') or '')


def _branding_logo_width(conn):
    """Width (px) the login page renders the branding logo at. Clamped to [40, 600]."""
    cfg = _get_setting(conn, 'branding', {}) or {}
    try:
        w = int(cfg.get('logo_width', 180) or 180)
    except (TypeError, ValueError):
        w = 180
    return max(40, min(600, w))


@app.route('/branding/logo')
def serve_branding_logo():
    """Public — the login page needs it before the user is authenticated."""
    conn = get_db()
    fname = _branding_filename(conn)
    conn.close()
    if not fname or not _BRANDING_NAME_RE.match(fname):
        return '', 404
    resp = send_from_directory(BRANDING_DIR, fname)
    # Neuter any scripting even if a rogue asset slipped in. Restrictive CSP
    # applies to the response itself when an SVG is browsed directly.
    resp.headers['Content-Security-Policy'] = "default-src 'none'; img-src 'self'; style-src 'unsafe-inline'"
    if fname.lower().endswith('.svg'):
        resp.headers['Content-Type'] = 'image/svg+xml'
    return resp


@app.route('/api/settings/branding', methods=['GET'])
@login_required
def get_branding_settings():
    conn = get_db()
    fname = _branding_filename(conn)
    width = _branding_logo_width(conn)
    conn.close()
    return jsonify({
        'logo_filename': fname,
        'logo_url': '/branding/logo' if fname else '',
        'logo_width': width,
    })


@app.route('/api/settings/branding', methods=['PUT'])
@admin_required
def update_branding_settings():
    """Update branding metadata — currently just the login-page logo width."""
    data = request.get_json() or {}
    conn = get_db()
    cfg = _get_setting(conn, 'branding', {}) or {}
    if 'logo_width' in data:
        try:
            w = int(data.get('logo_width'))
        except (TypeError, ValueError):
            conn.close()
            return jsonify({'error': 'logo_width must be a number'}), 400
        cfg['logo_width'] = max(40, min(600, w))
    _set_setting(conn, 'branding', cfg)
    conn.commit()
    new_width = _branding_logo_width(conn)
    conn.close()
    return jsonify({'logo_width': new_width})


def _sanitize_svg(svg_bytes):
    """Strip XSS vectors from an SVG: <script>, foreignObject, event handlers,
    `javascript:` URLs, and external href/xlink:href. Returns cleaned bytes
    or None if parsing fails. Admin-only upload path so this is belt-and-
    suspenders on top of the admin trust boundary."""
    import xml.etree.ElementTree as ET
    try:
        text = svg_bytes.decode('utf-8', errors='replace')
    except Exception:
        return None
    # Register SVG namespace so ElementTree round-trips element names cleanly
    ET.register_namespace('', 'http://www.w3.org/2000/svg')
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return None
    # Root must actually be an <svg>
    if not root.tag.lower().endswith('svg'):
        return None
    dangerous_tags = {'script', 'foreignobject', 'iframe', 'object', 'embed',
                      'video', 'audio', 'animate', 'animatetransform',
                      'animatemotion', 'set', 'handler', 'use'}

    def _local(tag):
        return tag.split('}', 1)[-1].lower()

    def _walk(el):
        # Iterate children list copy so we can remove safely
        for child in list(el):
            if _local(child.tag) in dangerous_tags:
                el.remove(child)
                continue
            # Strip unsafe attributes
            for attr in list(child.attrib.keys()):
                local_attr = attr.split('}', 1)[-1].lower()
                val = (child.attrib.get(attr) or '').strip().lower()
                if local_attr.startswith('on'):
                    del child.attrib[attr]
                    continue
                if local_attr in ('href', 'xlink:href') or local_attr.endswith(':href'):
                    if val.startswith(('javascript:', 'data:', 'vbscript:', 'file:')):
                        del child.attrib[attr]
                        continue
                if local_attr == 'style' and ('expression(' in val or 'javascript:' in val):
                    del child.attrib[attr]
            _walk(child)

    # Also strip attributes on the root element itself
    for attr in list(root.attrib.keys()):
        local_attr = attr.split('}', 1)[-1].lower()
        if local_attr.startswith('on'):
            del root.attrib[attr]
    _walk(root)
    try:
        cleaned = ET.tostring(root, encoding='utf-8', xml_declaration=True)
    except Exception:
        return None
    return cleaned


@app.route('/api/settings/branding/logo', methods=['POST'])
@admin_required
def upload_branding_logo():
    f = request.files.get('logo')
    if not f or not f.filename:
        return jsonify({'error': 'No file uploaded'}), 400
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in BRANDING_EXTS:
        return jsonify({'error': 'Logo must be a PNG or SVG'}), 400

    buf = f.read()
    new_fname = None
    save_bytes = None

    if ext == 'png':
        # Validate via Pillow + re-encode to strip metadata/chunks
        from PIL import Image, UnidentifiedImageError
        from io import BytesIO
        try:
            img = Image.open(BytesIO(buf))
            img.verify()
        except (UnidentifiedImageError, Exception):
            return jsonify({'error': 'Uploaded file is not a valid PNG image'}), 400
        img = Image.open(BytesIO(buf))
        if img.format != 'PNG':
            return jsonify({'error': 'Uploaded file is not a PNG'}), 400
        new_fname = f"logo-{uuid.uuid4().hex}.png"
        out = BytesIO()
        img.save(out, format='PNG', optimize=True)
        save_bytes = out.getvalue()
    else:  # svg
        cleaned = _sanitize_svg(buf)
        if cleaned is None:
            return jsonify({'error': 'Invalid or unsafe SVG (scripts and external refs are not allowed)'}), 400
        new_fname = f"logo-{uuid.uuid4().hex}.svg"
        save_bytes = cleaned

    conn = get_db()
    cfg = _get_setting(conn, 'branding', {}) or {}
    old = str(cfg.get('logo_filename', '') or '')
    out_path = os.path.join(BRANDING_DIR, new_fname)
    with open(out_path, 'wb') as fh:
        fh.write(save_bytes)
    # Remove the previous logo file (if any) so old uploads don't accumulate
    if old:
        try:
            prev = os.path.join(BRANDING_DIR, old)
            if os.path.exists(prev):
                os.remove(prev)
        except Exception:
            pass
    cfg['logo_filename'] = new_fname
    _set_setting(conn, 'branding', cfg)
    conn.commit()
    width = _branding_logo_width(conn)
    conn.close()
    return jsonify({'logo_filename': new_fname, 'logo_url': '/branding/logo', 'logo_width': width})


@app.route('/api/settings/branding/logo', methods=['DELETE'])
@admin_required
def delete_branding_logo():
    conn = get_db()
    cfg = _get_setting(conn, 'branding', {}) or {}
    old = str(cfg.get('logo_filename', '') or '')
    if old:
        try:
            prev = os.path.join(BRANDING_DIR, old)
            if os.path.exists(prev):
                os.remove(prev)
        except Exception:
            pass
    cfg['logo_filename'] = ''
    _set_setting(conn, 'branding', cfg)
    conn.commit()
    conn.close()
    return jsonify({'logo_filename': '', 'logo_url': ''})


@app.route('/api/settings/smtp/test', methods=['POST'])
@admin_required
def test_smtp():
    data = request.get_json() or {}
    to_email = str(data.get('to', '')).strip()
    if not to_email:
        return jsonify({'error': 'Recipient email required'}), 400
    ok, err = _send_email(
        to_email,
        'Warehouse Manager — SMTP Test',
        'This is a test email from Warehouse Manager. If you received this, SMTP is working.'
    )
    if ok:
        return jsonify({'success': True})
    return jsonify({'error': err or 'Send failed'}), 500


#
# Notification toggles — global (admin) + per-user (self)
# ─────────────────────────────────────────────────────────────

@app.route('/api/settings/notifications', methods=['GET'])
@login_required
def get_notifications_settings():
    """Global + current-user notification state. All users can read so the
    work-orders page knows whether to show the global-off banner and whether
    the user's personal toggle is on."""
    conn = get_db()
    global_enabled = bool(_get_setting(conn, 'smtp_notifications_enabled', True))
    row = conn.execute(
        "SELECT email_notifications_enabled FROM users WHERE id = ?",
        (current_user.id,)
    ).fetchone()
    conn.close()
    user_enabled = bool(row['email_notifications_enabled']) if row else True
    return jsonify({
        'global_enabled': global_enabled,
        'user_enabled': user_enabled,
    })


@app.route('/api/settings/notifications/global', methods=['PUT'])
@admin_required
def update_global_notifications():
    data = request.get_json() or {}
    enabled = bool(data.get('enabled', True))
    conn = get_db()
    _set_setting(conn, 'smtp_notifications_enabled', enabled)
    conn.commit()
    conn.close()
    return jsonify({'global_enabled': enabled})


@app.route('/api/settings/notifications/me', methods=['PUT'])
@login_required
def update_my_notifications():
    data = request.get_json() or {}
    enabled = 1 if bool(data.get('enabled', True)) else 0
    conn = get_db()
    conn.execute(
        "UPDATE users SET email_notifications_enabled = ? WHERE id = ?",
        (enabled, current_user.id)
    )
    conn.commit()
    conn.close()
    return jsonify({'user_enabled': bool(enabled)})


# ══════════════════════════════════════════
#  EMAIL (SMTP)
# ══════════════════════════════════════════

def _send_email(to_email, subject, body, attachments=None):
    """Send an email via configured SMTP. Returns (success, error_msg).
    attachments: optional list of dicts {filename, content (bytes), mime_type (e.g. 'image/jpeg')}.

    Honors two opt-outs: the global `smtp_notifications_enabled` setting
    (admin toggle) and the per-user `email_notifications_enabled` flag on the
    recipient's user record (matched by username = email)."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders
    from email.utils import formataddr

    if not to_email:
        return False, 'no recipient'

    conn = get_db()
    cfg = _get_setting(conn, 'smtp_config', {})
    global_enabled = _get_setting(conn, 'smtp_notifications_enabled', True)
    # Per-user opt-out: if the recipient has a matching user record with the
    # flag turned off, skip silently.
    user_row = conn.execute(
        "SELECT email_notifications_enabled FROM users WHERE LOWER(username) = LOWER(?)",
        (to_email,)
    ).fetchone()
    conn.close()

    if not bool(global_enabled):
        return False, 'notifications_disabled'
    if user_row is not None and 'email_notifications_enabled' in user_row.keys() \
            and not user_row['email_notifications_enabled']:
        return False, 'recipient_opted_out'

    # Defense-in-depth: strip CR/LF from anything going into a header so
    # header injection can't smuggle additional recipients or headers even
    # if Python's email module guardrails ever change.
    def _no_crlf(s):
        return str(s or '').replace('\r', ' ').replace('\n', ' ').strip()
    to_email = _no_crlf(to_email)
    subject = _no_crlf(subject)
    if not to_email or not _EMAIL_RE.match(to_email):
        return False, 'invalid recipient'

    host = cfg.get('host', '')
    if not host:
        return False, 'SMTP not configured'

    port = int(cfg.get('port', 587) or 587)
    username = cfg.get('username', '')
    password = cfg.get('password', '')
    use_tls = bool(cfg.get('use_tls', True))
    from_email = cfg.get('from_email', '') or username
    from_name = cfg.get('from_name', 'Warehouse Manager')

    if not from_email:
        return False, 'SMTP from_email not configured'

    try:
        if attachments:
            msg = MIMEMultipart()
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            for a in attachments:
                mime_type = a.get('mime_type', 'application/octet-stream')
                maintype, _, subtype = mime_type.partition('/')
                subtype = subtype or 'octet-stream'
                part = MIMEBase(maintype, subtype)
                part.set_payload(a.get('content', b''))
                encoders.encode_base64(part)
                fname = a.get('filename', 'attachment')
                part.add_header('Content-Disposition', f'attachment; filename="{fname}"')
                msg.attach(part)
        else:
            msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = subject
        msg['From'] = formataddr((from_name, from_email))
        msg['To'] = to_email

        if port == 465:
            server = smtplib.SMTP_SSL(host, port, timeout=15)
        else:
            server = smtplib.SMTP(host, port, timeout=15)
            if use_tls:
                server.starttls()
        if username:
            server.login(username, password)
        server.sendmail(from_email, [to_email], msg.as_string())
        server.quit()
        return True, None
    except Exception as e:
        app.logger.warning(f"SMTP send failed: {e}")
        return False, str(e)


def _format_parts_for_email(parts_json):
    import json as json_mod
    try:
        parts = json_mod.loads(parts_json or '[]') or []
    except Exception:
        parts = []
    if not parts:
        return ''
    lines = ['\nParts Requested:']
    for p in parts:
        lines.append(f"  - {p.get('quantity', 1)} × {p.get('description', '')}")
        d = (p.get('details') or '').strip()
        if d:
            lines.append(f"      {d}")
    return '\n'.join(lines) + '\n'


def _lookup_salesperson_email(conn, name):
    """Resolve a salesperson's display name (as stored on work_orders.sales_person)
    to their email, which is their user username. Users with is_sales_person=1 win
    first; falls back to any active user, since legacy WOs were created before the
    flag existed. Matches on display_name first, then username."""
    if not name:
        return ''
    n = str(name).strip()
    # Prefer flagged sales people
    row = conn.execute(
        "SELECT username FROM users WHERE active = 1 AND is_sales_person = 1 "
        "AND (LOWER(display_name) = LOWER(?) OR LOWER(username) = LOWER(?)) LIMIT 1",
        (n, n)
    ).fetchone()
    if row:
        return row['username']
    # Fallback: any active user by name (legacy records may reference a user that
    # hasn't been flagged as a salesperson yet)
    row = conn.execute(
        "SELECT username FROM users WHERE active = 1 "
        "AND (LOWER(display_name) = LOWER(?) OR LOWER(username) = LOWER(?)) LIMIT 1",
        (n, n)
    ).fetchone()
    if row:
        return row['username']
    return ''


# ══════════════════════════════════════════
#  WORK ORDERS API
# ══════════════════════════════════════════

def _compute_archive_after():
    """Return an ISO datetime string representing today's 23:00 local time.
    Used when a work order is marked delivered to schedule its auto-archival."""
    from datetime import datetime, time
    today_11pm = datetime.now().replace(hour=23, minute=0, second=0, microsecond=0)
    return today_11pm.strftime('%Y-%m-%d %H:%M:%S')


def _auto_archive_sweep(conn):
    """Archive delivered work orders whose archive_after threshold has passed.
    Called lazily at the top of list/count queries so no background process is needed."""
    from datetime import datetime
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute(
        "UPDATE work_orders SET archived_at = CURRENT_TIMESTAMP, was_archived = 1 "
        "WHERE status = 'delivered' AND archived_at IS NULL "
        "AND archive_after IS NOT NULL AND archive_after <= ?",
        (now_str,)
    )


def _assign_wo_number(conn):
    """Generate next work-order number: WO-00001 ... WO-99999, then WO-100000+
    Tolerant of both old (WO#####) and new (WO-#####) stored formats when scanning for max."""
    rows = conn.execute("SELECT wo_number FROM work_orders WHERE wo_number LIKE 'WO%'").fetchall()
    max_num = 0
    for r in rows:
        s = (r['wo_number'] or '').replace('WO-', '').replace('WO', '')
        try:
            n = int(s)
            if n > max_num:
                max_num = n
        except ValueError:
            continue
    next_num = max_num + 1
    if next_num < 100000:
        return f"WO-{next_num:05d}"
    return f"WO-{next_num}"


def _work_order_to_dict(conn, row):
    import json as json_mod
    d = dict(row)
    notes_rows = conn.execute(
        "SELECT id, note, author, author_user_id, parent_id, note_type, created_at "
        "FROM work_order_notes WHERE work_order_id = ? "
        "ORDER BY created_at DESC, id DESC",
        (d['id'],)
    ).fetchall()
    d['notes_log'] = [dict(n) for n in notes_rows]
    try:
        raw_parts = json_mod.loads(d.get('parts_json') or '[]') or []
    except Exception:
        raw_parts = []
    # Backfill per-part state fields so legacy rows have consistent shape
    parts = []
    mutated = False
    for p in raw_parts:
        if not isinstance(p, dict):
            continue
        key = str(p.get('key', '') or '').strip()
        if not key:
            key = uuid.uuid4().hex
            mutated = True
        parts.append({
            'key': key,
            'description': str(p.get('description', '')),
            'details': str(p.get('details', '') or ''),
            'quantity': int(p.get('quantity', 1) or 1),
            'pulled': bool(p.get('pulled', False)),
            'pulled_at': str(p.get('pulled_at', '') or ''),
            'flagged': bool(p.get('flagged', False)),
            'flag_note': str(p.get('flag_note', '') or ''),
        })
    # Persist any newly-generated keys so photos uploaded later stay anchored
    if mutated:
        conn.execute(
            "UPDATE work_orders SET parts_json = ? WHERE id = ?",
            (json_mod.dumps(parts), d['id'])
        )
        conn.commit()
        d['parts_json'] = json_mod.dumps(parts)

    # Attach per-part photos
    if parts:
        keys = [p['key'] for p in parts]
        placeholders = ','.join(['?'] * len(keys))
        photo_rows = conn.execute(
            f"SELECT id, part_key, filename, comment, uploaded_by, created_at "
            f"FROM work_order_part_photos "
            f"WHERE work_order_id = ? AND part_key IN ({placeholders}) "
            f"ORDER BY created_at, id",
            [d['id']] + keys
        ).fetchall()
        by_key = {}
        for r in photo_rows:
            by_key.setdefault(r['part_key'], []).append(dict(r))
        for p in parts:
            p['photos'] = by_key.get(p['key'], [])
    d['parts'] = parts
    return d


def _normalize_parts(raw):
    """Coerce a client-supplied parts list into validated
    [{key, description, details, quantity, pulled, flagged, flag_note}, ...].
    Preserves a stable per-part key (generated if absent) so photos
    keyed off it survive edits and reordering."""
    out = []
    for p in (raw or []):
        if not isinstance(p, dict):
            continue
        desc = str(p.get('description', '')).strip()
        details = str(p.get('details', '') or '').strip()
        try:
            qty = int(p.get('quantity', 1) or 1)
        except (TypeError, ValueError):
            qty = 1
        if qty < 1:
            qty = 1
        if desc:
            key = str(p.get('key', '') or '').strip() or uuid.uuid4().hex
            out.append({
                'key': key,
                'description': desc,
                'details': details,
                'quantity': qty,
                'pulled': bool(p.get('pulled', False)),
                'pulled_at': str(p.get('pulled_at', '') or ''),
                'flagged': bool(p.get('flagged', False)),
                'flag_note': str(p.get('flag_note', '') or '').strip(),
            })
    return out


def _actor():
    return (current_user.display_name or current_user.username) if current_user.is_authenticated else 'system'


def _log_wo_audit(conn, wid, action, description):
    """Record a single audit entry. Caller owns the transaction (we do not commit)."""
    conn.execute(
        "INSERT INTO work_order_audit (work_order_id, action, actor, description) VALUES (?, ?, ?, ?)",
        (wid, action, _actor(), description)
    )


_WO_FIELD_LABELS = {
    'warehouse_location': 'Warehouse Location',
    'customer_name': 'Customer',
    'quote_invoice': 'Quote/Invoice',
    'sales_person': 'Sales Person',
    'vehicle': 'Vehicle',
    'vin': 'VIN',
    'priority': 'Priority',
    'notes': 'Request Details',
}


def _diff_wo_fields(old_row, new_data):
    """Return a list of human-readable change strings for fields that changed."""
    changes = []
    for key, label in _WO_FIELD_LABELS.items():
        if key not in new_data:
            continue
        old = (old_row[key] or '') if key in old_row.keys() else ''
        new = str(new_data.get(key, '') or '').strip()
        if str(old).strip() == new:
            continue
        if key == 'notes':
            # Request details can be long — just say "updated"
            changes.append(f"{label} updated")
        else:
            changes.append(f"{label}: '{old or '—'}' → '{new or '—'}'")
    return changes


def _diff_parts(old_json, new_list):
    """Return a brief summary of parts-list changes, or empty string if unchanged."""
    import json as json_mod
    try:
        old_parts = json_mod.loads(old_json or '[]') or []
    except Exception:
        old_parts = []
    # Normalize for comparison
    norm = lambda ps: [(p.get('quantity', 1), (p.get('description') or '').strip()) for p in ps]
    if norm(old_parts) == norm(new_list):
        return ''
    return f"Parts list updated ({len(old_parts)} → {len(new_list)} item{'s' if len(new_list) != 1 else ''})"


@app.route('/api/work-orders', methods=['GET'])
@login_required
def list_work_orders():
    """List work orders. Optional ?status=requested,flagged,delivered or ?archived=1"""
    status_filter = request.args.get('status', '')
    archived = request.args.get('archived', '') == '1'

    conn = get_db()
    _auto_archive_sweep(conn)
    conds, params = [], []
    if archived:
        conds.append("archived_at IS NOT NULL")
    elif status_filter:
        wanted = [s.strip() for s in status_filter.split(',') if s.strip()]
        if wanted:
            placeholders = ','.join(['?'] * len(wanted))
            conds.append(f"status IN ({placeholders})")
            params.extend(wanted)
        conds.append("archived_at IS NULL")
    else:
        # default: active list = anything not yet archived (including pending-delivered)
        conds.append("archived_at IS NULL")

    where = "WHERE " + " AND ".join(conds) if conds else ""
    search = request.args.get('search', '').strip()
    if search:
        term = f"%{search}%"
        search_cols = ['wo_number', 'customer_name', 'quote_invoice', 'sales_person',
                       'vehicle', 'vin', 'warehouse_location', 'notes', 'parts_json']
        where = (where + " AND " if where else "WHERE ") + "(" + " OR ".join(f"{c} LIKE ?" for c in search_cols) + ")"
        params.extend([term] * len(search_cols))

    SORT_ALLOWED = {'request_date', 'customer_name', 'priority', 'wo_number', 'status'}
    sort_by = request.args.get('sort_by', 'request_date')
    sort_dir = request.args.get('sort_dir', 'desc').lower()
    if sort_by not in SORT_ALLOWED:
        sort_by = 'request_date'
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'desc'

    rows = conn.execute(
        f"SELECT * FROM work_orders {where} ORDER BY {sort_by} COLLATE NOCASE {sort_dir.upper()}, id {sort_dir.upper()}",
        params
    ).fetchall()
    result = [_work_order_to_dict(conn, r) for r in rows]
    conn.close()
    return jsonify(result)


@app.route('/api/work-orders/counts', methods=['GET'])
@login_required
def work_order_counts():
    conn = get_db()
    _auto_archive_sweep(conn)
    counts = {
        'requested':  conn.execute("SELECT COUNT(*) FROM work_orders WHERE status='requested' AND archived_at IS NULL").fetchone()[0],
        'flagged':    conn.execute("SELECT COUNT(*) FROM work_orders WHERE status='flagged'   AND archived_at IS NULL").fetchone()[0],
        # Delivered but still in active (awaiting auto-archive or manual Archive Now)
        'delivered_pending': conn.execute("SELECT COUNT(*) FROM work_orders WHERE status='delivered' AND archived_at IS NULL").fetchone()[0],
        # Archived (what shows in the Archive view)
        'delivered':  conn.execute("SELECT COUNT(*) FROM work_orders WHERE archived_at IS NOT NULL").fetchone()[0],
    }
    counts['active'] = counts['requested'] + counts['flagged'] + counts['delivered_pending']
    conn.close()
    return jsonify(counts)


@app.route('/api/work-orders/<int:wid>', methods=['GET'])
@login_required
def get_work_order(wid):
    conn = get_db()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    wo = _work_order_to_dict(conn, row)
    conn.close()
    return jsonify(wo)


@app.route('/api/work-orders', methods=['POST'])
@editor_required
def create_work_order():
    import json as json_mod
    data = request.get_json() or {}
    conn = get_db()
    wo_num = _assign_wo_number(conn)
    created_by = current_user.display_name or current_user.username
    parts = _normalize_parts(data.get('parts'))
    cur = conn.execute('''
        INSERT INTO work_orders
            (wo_number, warehouse_location, customer_name, quote_invoice, sales_person,
             vehicle, vin, priority, notes, status, created_by, created_by_user_id, parts_json)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        wo_num,
        str(data.get('warehouse_location', '')).strip(),
        str(data.get('customer_name', '')).strip(),
        str(data.get('quote_invoice', '')).strip(),
        str(data.get('sales_person', '')).strip(),
        str(data.get('vehicle', '')).strip(),
        str(data.get('vin', '')).strip(),
        str(data.get('priority', 'Normal')).strip() or 'Normal',
        str(data.get('notes', '')).strip(),
        'requested',
        created_by,
        current_user.id,
        json_mod.dumps(parts),
    ))
    wid = cur.lastrowid
    _log_wo_audit(conn, wid, 'created', f"Work order created as {wo_num}")
    conn.commit()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, row)
    conn.close()
    return jsonify(wo), 201


@app.route('/api/work-orders/<int:wid>/archive-now', methods=['POST'])
@editor_required
def archive_now_work_order(wid):
    """Manually archive a delivered work order immediately (bypass the 23:00 sweep)."""
    conn = get_db()
    row = conn.execute("SELECT id, status, archived_at, wo_number FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if row['status'] != 'delivered':
        conn.close()
        return jsonify({'error': 'Only delivered work orders can be archived'}), 400
    if row['archived_at']:
        conn.close()
        return jsonify({'error': 'Work order is already archived'}), 400
    conn.execute(
        "UPDATE work_orders SET archived_at = CURRENT_TIMESTAMP, was_archived = 1 WHERE id = ?",
        (wid,)
    )
    _log_wo_audit(conn, wid, 'status_changed', "Manually archived")
    conn.execute(
        "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'general')",
        (wid, "Manually archived.", _actor())
    )
    conn.commit()
    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify(wo)


@app.route('/api/work-orders/<int:wid>/archive', methods=['POST'])
@editor_required
def archive_work_order(wid):
    """Archive a work order in *any* status (requested, flagged, delivered).
    Non-delivered archives keep was_delivered=0 so the archive view can flag
    them 'Not Delivered'. Reopening (via the existing status endpoint) will
    put it back to requested/flagged with was_delivered still at 0."""
    conn = get_db()
    row = conn.execute("SELECT id, status, archived_at, wo_number FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if row['archived_at']:
        conn.close()
        return jsonify({'error': 'Work order is already archived'}), 400
    conn.execute(
        "UPDATE work_orders SET archived_at = CURRENT_TIMESTAMP, was_archived = 1, "
        "updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (wid,)
    )
    desc = "Archived" if row['status'] == 'delivered' else f"Archived (status: {row['status']}, not delivered)"
    _log_wo_audit(conn, wid, 'status_changed', desc)
    conn.execute(
        "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'general')",
        (wid, desc + '.', _actor())
    )
    conn.commit()
    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify(wo)


@app.route('/api/work-orders/<int:wid>/duplicate', methods=['POST'])
@editor_required
def duplicate_work_order(wid):
    """Create a new work order by copying fields + parts from an existing one.
    Status resets to 'requested'; per-part pulled/flagged state is reset;
    notes_log and audit trail are fresh (audit records the source WO #)."""
    import json as json_mod
    conn = get_db()
    src = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not src:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    try:
        src_parts = json_mod.loads(src['parts_json'] or '[]') or []
    except Exception:
        src_parts = []
    fresh_parts = []
    for p in src_parts:
        if not isinstance(p, dict):
            continue
        desc = str(p.get('description', '')).strip()
        if not desc:
            continue
        try:
            qty = int(p.get('quantity', 1) or 1)
        except (TypeError, ValueError):
            qty = 1
        if qty < 1:
            qty = 1
        fresh_parts.append({
            'description': desc,
            'details': str(p.get('details', '') or ''),
            'quantity': qty,
            'pulled': False,
            'pulled_at': '',
            'flagged': False,
            'flag_note': '',
        })

    wo_num = _assign_wo_number(conn)
    created_by = _actor()
    cur = conn.execute('''
        INSERT INTO work_orders
            (wo_number, warehouse_location, customer_name, quote_invoice, sales_person,
             vehicle, vin, priority, notes, status, created_by, created_by_user_id, parts_json)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        wo_num,
        src['warehouse_location'] or '',
        src['customer_name'] or '',
        src['quote_invoice'] or '',
        src['sales_person'] or '',
        src['vehicle'] or '',
        src['vin'] or '',
        src['priority'] or 'Normal',
        src['notes'] or '',
        'requested',
        created_by,
        (current_user.id if current_user.is_authenticated else None),
        json_mod.dumps(fresh_parts),
    ))
    new_id = cur.lastrowid
    _log_wo_audit(conn, new_id, 'created',
                  f"Created {wo_num} by duplicating {src['wo_number']}")
    # Also leave a visible note on the new WO's activity feed for context
    conn.execute(
        "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'general')",
        (new_id, f"Duplicated from {src['wo_number']}.", created_by)
    )
    conn.commit()

    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (new_id,)).fetchone()
    wo = _work_order_to_dict(conn, row)
    conn.close()
    return jsonify(wo), 201


@app.route('/api/work-orders/<int:wid>', methods=['PUT'])
@editor_required
def update_work_order(wid):
    data = request.get_json() or {}
    conn = get_db()
    existing = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    if existing['status'] == 'delivered':
        conn.close()
        return jsonify({'error': 'Delivered work orders cannot be edited. Reopen first.'}), 400
    # Editing the original request body is restricted to the originator, admins,
    # and supervisors. Plain editors who didn't create this WO can still add
    # notes, upload photos, mark parts pulled/flagged — just not rewrite the
    # request fields or the parts list.
    is_originator = existing['created_by_user_id'] == current_user.id \
        if 'created_by_user_id' in existing.keys() else False
    if not (current_user.is_admin or current_user.role == 'supervisor' or is_originator):
        conn.close()
        return jsonify({
            'error': 'Only the originator, an admin, or a supervisor can edit this work order. You can still add notes, photos, and flag parts.'
        }), 403

    import json as json_mod
    editable = ['warehouse_location', 'customer_name', 'quote_invoice', 'sales_person',
                'vehicle', 'vin', 'priority', 'notes']
    # Compute diff before applying
    change_parts = _diff_wo_fields(existing, data)
    new_parts_list = None
    if 'parts' in data:
        new_parts_list = _normalize_parts(data.get('parts'))
        parts_diff = _diff_parts(existing['parts_json'] if 'parts_json' in existing.keys() else '[]', new_parts_list)
        if parts_diff:
            change_parts.append(parts_diff)

    sets, vals = [], []
    for k in editable:
        if k in data:
            sets.append(f"{k}=?")
            vals.append(str(data.get(k, '')).strip())
    if new_parts_list is not None:
        sets.append("parts_json=?")
        vals.append(json_mod.dumps(new_parts_list))
    if sets:
        sets.append("updated_at=CURRENT_TIMESTAMP")
        vals.append(wid)
        conn.execute(f"UPDATE work_orders SET {', '.join(sets)} WHERE id=?", vals)
        # If parts were replaced, drop photos tied to part keys that no longer exist
        if new_parts_list is not None:
            keep_keys = {p.get('key') for p in new_parts_list if p.get('key')}
            orphans = conn.execute(
                "SELECT id, part_key, filename FROM work_order_part_photos WHERE work_order_id = ?",
                (wid,)
            ).fetchall()
            for orp in orphans:
                if orp['part_key'] not in keep_keys:
                    delete_image(orp['filename'])
                    conn.execute("DELETE FROM work_order_part_photos WHERE id = ?", (orp['id'],))
        if change_parts:
            _log_wo_audit(conn, wid, 'edited', '; '.join(change_parts))
        conn.commit()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, row)
    conn.close()
    return jsonify(wo)


@app.route('/api/work-orders/<int:wid>/status', methods=['POST'])
@editor_required
def set_work_order_status(wid):
    data = request.get_json() or {}
    new_status = str(data.get('status', '')).strip().lower()
    # 'flagged' is no longer directly settable — it's derived from per-part flags.
    if new_status not in ('requested', 'delivered'):
        return jsonify({'error': 'Invalid status (flag status is derived from per-part flags)'}), 400

    conn = get_db()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    old_status = row['status']

    if new_status == 'delivered':
        # Schedule auto-archival for 23:00 local time on the day of delivery.
        # archived_at stays NULL so the WO stays in Active until the sweep (or the
        # Archive Now button) flips it. was_delivered latches ON so delete gating
        # can tell "reopened-from-delivered" apart from "reopened-from-archive".
        conn.execute(
            "UPDATE work_orders SET status = ?, completed_at = CURRENT_TIMESTAMP, "
            "archive_after = ?, archived_at = NULL, was_delivered = 1, "
            "updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (new_status, _compute_archive_after(), wid)
        )
    else:
        # Reopening → clear delivered/archive state, then recompute (flag vs requested)
        # from per-part flags.
        conn.execute(
            "UPDATE work_orders SET status = ?, completed_at = NULL, archive_after = NULL, "
            "archived_at = NULL, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            ('requested', wid)
        )
        import json as json_mod
        try:
            parts = json_mod.loads(row['parts_json'] or '[]') or []
        except Exception:
            parts = []
        new_status = _recompute_wo_status(conn, wid, parts, 'requested')

    # Audit log the status change + mirror into the Notes & Activity feed
    if new_status != old_status:
        actor = _actor()
        if new_status == 'delivered':
            audit_desc = f"Status: {old_status} → delivered"
            conn.execute(
                "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'general')",
                (wid, "Marked as delivered.", actor)
            )
        else:
            if old_status == 'delivered':
                reason = f"Reopened from delivered — now {new_status}."
            else:
                reason = f"Status changed {old_status} → {new_status}."
            audit_desc = f"Status: {old_status} → {new_status} (reopened)"
            conn.execute(
                "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'general')",
                (wid, reason, actor)
            )
        _log_wo_audit(conn, wid, 'status_changed', audit_desc)

    conn.commit()

    # Send status-change email to salesperson on delivery (flag emails come from per-part path)
    email_sent = False
    email_error = None
    if new_status == 'delivered' and new_status != old_status:
        sp_email = _lookup_salesperson_email(conn, row['sales_person'])
        if sp_email:
            wo_num = row['wo_number']
            customer = row['customer_name']
            parts_block = _format_parts_for_email(row['parts_json'] if 'parts_json' in row.keys() else '[]')
            subject = f"[Delivered] Work Order {wo_num}"
            body = (
                f"Work Order: {wo_num}\n"
                f"Customer: {customer}\n"
                f"Vehicle: {row['vehicle']}\n"
                f"Status: DELIVERED / COMPLETE\n"
                f"{parts_block}"
            )
            email_sent, email_error = _send_email(sp_email, subject, body)

    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify({'work_order': wo, 'email_sent': email_sent, 'email_error': email_error})


@app.route('/api/work-orders/<int:wid>/notes', methods=['POST'])
@editor_required
def add_work_order_note(wid):
    data = request.get_json() or {}
    note = str(data.get('note', '')).strip()
    if not note:
        return jsonify({'error': 'Note is required'}), 400

    conn = get_db()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    author = current_user.display_name or current_user.username
    conn.execute(
        "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'flag')",
        (wid, note, author)
    )
    # Keep flag_note mirrored with the latest note
    conn.execute("UPDATE work_orders SET flag_note = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (note, wid))
    _log_wo_audit(conn, wid, 'note_added', f"Flag note added: {note}")
    conn.commit()

    # If work order is flagged, send an email update on every new note
    email_sent = False
    email_error = None
    if row['status'] == 'flagged':
        sp_email = _lookup_salesperson_email(conn, row['sales_person'])
        if sp_email:
            subject = f"[Flagged — Update] Work Order {row['wo_number']}"
            body = (
                f"Work Order: {row['wo_number']}\n"
                f"Customer: {row['customer_name']}\n"
                f"Vehicle: {row['vehicle']}\n"
                f"Status: FLAGGED (new update)\n\n"
                f"New note from {author}:\n{note}\n"
            )
            email_sent, email_error = _send_email(sp_email, subject, body)

    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify({'work_order': wo, 'email_sent': email_sent, 'email_error': email_error})


def _collect_note_recipients(conn, wo_row, parent_id, current_email):
    """Return a sorted list of email recipients for a new note or reply.
    Always includes the WO's salesperson email + every unique user who has
    already posted in the thread. The current author is excluded."""
    emails = []
    seen = set()

    def add(email):
        if not email:
            return
        key = email.strip().lower()
        if not key or key == (current_email or '').strip().lower() or key in seen:
            return
        seen.add(key)
        emails.append(email.strip())

    # Salesperson on the WO
    sp_email = _lookup_salesperson_email(conn, wo_row['sales_person']) if wo_row else ''
    add(sp_email)

    # Prior participants in this thread (only matters for replies)
    if parent_id is not None:
        rows = conn.execute(
            "SELECT u.username AS email FROM work_order_notes n "
            "LEFT JOIN users u ON u.id = n.author_user_id "
            "WHERE n.work_order_id = ? AND (n.id = ? OR n.parent_id = ?) "
            "  AND u.username IS NOT NULL",
            (wo_row['id'], parent_id, parent_id)
        ).fetchall()
        for r in rows:
            add(r['email'])

    return emails


def _send_note_email(wo_row, author_name, note_text, parent_id, recipients, is_reply):
    """Fire off the notification email for a new note / reply."""
    if not recipients:
        return False, 'no_recipients'
    wo_num = wo_row['wo_number']
    label = 'Reply' if is_reply else 'Note'
    subject = f"[{label}] Work Order {wo_num}"
    lines = [
        f"Work Order: {wo_num}",
        f"Customer: {wo_row['customer_name'] or '—'}",
        f"Vehicle: {wo_row['vehicle'] or '—'}",
        '',
        f"{author_name} added a {'reply' if is_reply else 'note'}:",
        '',
        note_text,
        '',
        'Log in to Warehouse Manager to reply.',
    ]
    body = '\n'.join(lines) + '\n'
    last_ok, last_err = False, None
    for to in recipients:
        ok, err = _send_email(to, subject, body)
        if ok:
            last_ok = True
        elif err:
            last_err = err
    return last_ok, last_err


@app.route('/api/work-orders/<int:wid>/general-notes', methods=['POST'])
@editor_required
def add_general_note(wid):
    """Append a running work-order note. Optionally `parent_id` threads it as a
    reply to an existing note. Sends a notification email to the salesperson +
    every prior participant in the thread (minus the current author)."""
    data = request.get_json() or {}
    note = str(data.get('note', '') or '').strip()
    if not note:
        return jsonify({'error': 'Note is required'}), 400

    parent_id = data.get('parent_id')
    try:
        parent_id = int(parent_id) if parent_id not in (None, '', 0, '0') else None
    except (TypeError, ValueError):
        parent_id = None

    conn = get_db()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    # If this is a reply, make sure the parent exists on this WO and resolve to
    # the root of the thread (flat one-level threading).
    if parent_id is not None:
        parent = conn.execute(
            "SELECT id, parent_id FROM work_order_notes WHERE id = ? AND work_order_id = ?",
            (parent_id, wid)
        ).fetchone()
        if not parent:
            conn.close()
            return jsonify({'error': 'Parent note not found'}), 404
        # Re-parent replies-to-replies to the root so the thread stays flat
        if parent['parent_id']:
            parent_id = parent['parent_id']

    author = current_user.display_name or current_user.username
    author_user_id = current_user.id
    conn.execute(
        "INSERT INTO work_order_notes "
        "(work_order_id, note, author, author_user_id, note_type, parent_id) "
        "VALUES (?, ?, ?, ?, 'general', ?)",
        (wid, note, author, author_user_id, parent_id)
    )
    conn.execute("UPDATE work_orders SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (wid,))
    _log_wo_audit(conn, wid, 'note_added', f"{'Reply' if parent_id else 'Note'}: {note}")
    conn.commit()

    # Determine recipients and send
    current_email = current_user.username
    recipients = _collect_note_recipients(conn, row, parent_id, current_email)
    email_sent, email_error = _send_note_email(
        row, author, note, parent_id, recipients, is_reply=(parent_id is not None)
    )

    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify({
        'work_order': wo,
        'email_sent': email_sent,
        'email_error': None if email_sent else (email_error or 'no_recipients'),
        'email_recipients': recipients,
    })


@app.route('/api/work-orders/<int:wid>', methods=['DELETE'])
@editor_required
def delete_work_order(wid):
    """Delete a work order. Gating:
      * Base rule: only the originator, admins, and supervisors may delete.
      * Reopened-from-delivered (was_delivered=1, currently active): NO delete
        for anyone — users can Archive instead.
      * Currently delivered (pending or archived): admin only.
    """
    conn = get_db()
    existing = conn.execute(
        "SELECT id, wo_number, status, was_archived, was_delivered, archived_at, "
        "created_by_user_id FROM work_orders WHERE id = ?", (wid,)
    ).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    is_admin = current_user.is_admin
    is_supervisor = current_user.role == 'supervisor'
    is_originator = existing['created_by_user_id'] == current_user.id
    is_currently_archived = bool(existing['archived_at'])
    is_delivered = existing['status'] == 'delivered'

    # Reopened-from-delivered lock: no one can delete, even admin.
    if (existing['was_delivered'] and not is_delivered and not is_currently_archived):
        conn.close()
        return jsonify({
            'error': 'Delivered work orders cannot be deleted once completed — use Archive instead.'
        }), 403

    if is_delivered or is_currently_archived:
        if not is_admin:
            conn.close()
            return jsonify({
                'error': 'Delivered / archived work orders can only be deleted by an admin.'
            }), 403
    else:
        if not (is_admin or is_supervisor or is_originator):
            conn.close()
            return jsonify({
                'error': 'Only the originator, an admin, or a supervisor can delete this work order.'
            }), 403
    _log_wo_audit(conn, wid, 'deleted', f"Work order {existing['wo_number']} deleted")
    # Clean up per-part photo files from disk before removing rows
    photo_rows = conn.execute(
        "SELECT filename FROM work_order_part_photos WHERE work_order_id = ?", (wid,)
    ).fetchall()
    for pr in photo_rows:
        delete_image(pr['filename'])
    conn.execute("DELETE FROM work_order_part_photos WHERE work_order_id = ?", (wid,))
    conn.execute("DELETE FROM work_order_notes WHERE work_order_id = ?", (wid,))
    conn.execute("DELETE FROM work_orders WHERE id = ?", (wid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/work-orders/<int:wid>/re-archive', methods=['POST'])
@editor_required
def re_archive_work_order(wid):
    """Immediately re-archive a work order (skipping the 23:00 grace period).
    Used from the UI where editors can't delete a previously-archived reopened WO."""
    conn = get_db()
    row = conn.execute("SELECT id, status, wo_number FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    conn.execute(
        "UPDATE work_orders SET status = 'delivered', completed_at = CURRENT_TIMESTAMP, "
        "archive_after = ?, archived_at = CURRENT_TIMESTAMP, was_archived = 1, was_delivered = 1, "
        "flag_note = '', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (_compute_archive_after(), wid)
    )
    _log_wo_audit(conn, wid, 'status_changed', "Re-archived")
    conn.execute(
        "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'general')",
        (wid, "Re-archived.", _actor())
    )
    conn.commit()
    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify(wo)


def _recompute_wo_status(conn, wid, parts, current_status):
    """Derive the WO's top-level status from its per-part flags.
    - delivered WOs stay delivered (per-part flags don't reopen them)
    - any part flagged → 'flagged'
    - otherwise → 'requested'
    Also keeps work_orders.flag_note in sync (cleared unless flagged, in which case
    it mirrors the first flagged part's note for quick display in email templates).
    Returns the new status string."""
    if current_status == 'delivered':
        return current_status
    any_flagged = any(isinstance(p, dict) and bool(p.get('flagged')) for p in parts)
    new_status = 'flagged' if any_flagged else 'requested'
    if any_flagged:
        first_flag_note = next(
            (str(p.get('flag_note', '') or '') for p in parts
             if isinstance(p, dict) and p.get('flagged')),
            ''
        )
        conn.execute(
            "UPDATE work_orders SET status = ?, flag_note = ? WHERE id = ?",
            (new_status, first_flag_note, wid)
        )
    else:
        conn.execute(
            "UPDATE work_orders SET status = ?, flag_note = '' WHERE id = ?",
            (new_status, wid)
        )
    return new_status


def _update_part_field(wid, idx, updates, audit_desc, email_subject=None, email_body_extra=None, history_note=None, history_note_type='flag'):
    """Shared helper: load WO, mutate parts[idx] with updates dict, persist, audit-log.
    If history_note is provided, also insert into work_order_notes so it shows up in
    the flag notes history (use history_note_type to pick 'flag' or 'general').
    After persisting, recomputes the WO's derived status from per-part flags.
    Returns (wo_dict, email_sent, email_error)."""
    import json as json_mod
    conn = get_db()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return None, False, 'not_found'
    try:
        parts = json_mod.loads(row['parts_json'] or '[]') or []
    except Exception:
        parts = []
    if idx < 0 or idx >= len(parts):
        conn.close()
        return None, False, 'bad_index'
    part = parts[idx]
    for k, v in updates.items():
        part[k] = v
    parts[idx] = part
    conn.execute(
        "UPDATE work_orders SET parts_json=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (json_mod.dumps(parts), wid)
    )
    _recompute_wo_status(conn, wid, parts, row['status'])
    if history_note:
        # Tag the note with the acting user so replies in the thread can email
        # them back. Flag notes live in the same thread as general notes so
        # teammates can reply to "Part flagged — …" directly.
        author_user_id = current_user.id if current_user.is_authenticated else None
        conn.execute(
            "INSERT INTO work_order_notes (work_order_id, note, author, author_user_id, note_type) "
            "VALUES (?, ?, ?, ?, ?)",
            (wid, history_note, _actor(), author_user_id, history_note_type)
        )
    _log_wo_audit(conn, wid, 'edited', audit_desc)
    conn.commit()

    email_sent, email_error = False, None
    if email_subject:
        sp_email = _lookup_salesperson_email(conn, row['sales_person'])
        if sp_email:
            body = (
                f"Work Order: {row['wo_number']}\n"
                f"Customer: {row['customer_name']}\n"
                f"Vehicle: {row['vehicle']}\n"
                f"{email_body_extra or ''}"
            )
            email_sent, email_error = _send_email(sp_email, email_subject, body)

    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return wo, email_sent, email_error


@app.route('/api/work-orders/<int:wid>/parts/<int:idx>/pulled', methods=['POST'])
@editor_required
def set_part_pulled(wid, idx):
    from datetime import datetime
    import json as json_mod
    data = request.get_json() or {}
    pulled = bool(data.get('pulled', True))
    pulled_at = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S') if pulled else ''

    # Pull the part description for a human-readable activity note
    conn = get_db()
    row = conn.execute("SELECT parts_json FROM work_orders WHERE id = ?", (wid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    try:
        parts = json_mod.loads(row['parts_json'] or '[]') or []
    except Exception:
        parts = []
    if idx < 0 or idx >= len(parts):
        return jsonify({'error': 'Invalid part index'}), 400
    part_desc = f"{parts[idx].get('quantity', 1)} × {parts[idx].get('description', '')}"

    if pulled:
        # Pulled: post a note + audit entry as usual
        wo, _, err = _update_part_field(
            wid, idx,
            {'pulled': pulled, 'pulled_at': pulled_at},
            audit_desc=f"Part {idx+1} marked pulled at {pulled_at}",
            history_note=f"Part pulled — {part_desc}",
            history_note_type='general',
        )
    else:
        # Unmarked-pulled: keep the audit entry, but clean up the notes thread.
        # If the most recent "Part pulled" note for this part has no replies,
        # remove it silently (the whole "pulled → un-pulled" ping-pong
        # disappears). If a teammate already replied to it, preserve the
        # original note and post an explicit "Part unmarked pulled" entry.
        c2 = get_db()
        expected = f"Part pulled — {part_desc}"
        last = c2.execute(
            "SELECT id FROM work_order_notes "
            "WHERE work_order_id = ? AND note = ? AND note_type = 'general' "
            "ORDER BY created_at DESC, id DESC LIMIT 1",
            (wid, expected)
        ).fetchone()
        has_replies = False
        if last:
            reply = c2.execute(
                "SELECT id FROM work_order_notes WHERE parent_id = ? LIMIT 1",
                (last['id'],)
            ).fetchone()
            has_replies = bool(reply)
        if last and not has_replies:
            # Remove the original pulled note; don't add a new history note.
            c2.execute("DELETE FROM work_order_notes WHERE id = ?", (last['id'],))
            c2.commit()
            c2.close()
            wo, _, err = _update_part_field(
                wid, idx,
                {'pulled': pulled, 'pulled_at': pulled_at},
                audit_desc=f"Part {idx+1} marked not pulled (pulled note removed from thread)",
                history_note=None,
            )
        else:
            c2.close()
            wo, _, err = _update_part_field(
                wid, idx,
                {'pulled': pulled, 'pulled_at': pulled_at},
                audit_desc=f"Part {idx+1} marked not pulled",
                history_note=f"Part unmarked pulled — {part_desc}",
                history_note_type='general',
            )

    if wo is None:
        return jsonify({'error': 'Not found' if err == 'not_found' else 'Invalid part index'}), 404 if err == 'not_found' else 400
    return jsonify(wo)


@app.route('/api/work-orders/<int:wid>/parts/<int:idx>/flag', methods=['POST'])
@editor_required
def set_part_flag(wid, idx):
    """Flag a part, update an existing flag's note, or unflag.
      - {flagged: true, flag_note: "..."}  → flag (or update note if already flagged)
      - {flagged: false}                   → unflag"""
    data = request.get_json() or {}
    flagged = bool(data.get('flagged', True))
    note = str(data.get('flag_note', '') or '').strip()
    if flagged and not note:
        return jsonify({'error': 'Flag note is required'}), 400

    # Fetch part description for audit/email context
    conn = get_db()
    row = conn.execute("SELECT parts_json FROM work_orders WHERE id = ?", (wid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    import json as json_mod
    try:
        parts = json_mod.loads(row['parts_json'] or '[]') or []
    except Exception:
        parts = []
    if idx < 0 or idx >= len(parts):
        return jsonify({'error': 'Invalid part index'}), 400
    part_desc = f"{parts[idx].get('quantity', 1)} × {parts[idx].get('description', '')}"
    was_flagged = bool(parts[idx].get('flagged'))
    old_note = str(parts[idx].get('flag_note', '') or '')

    # History-note type defaults to 'flag' (which styles red in the thread).
    # Unflag events are neutral activity, so they use 'general'.
    history_note_type = 'flag'
    if flagged:
        updates = {'flagged': True, 'flag_note': note}
        if was_flagged and old_note == note:
            # Idempotent update — return current state without writing
            c2 = get_db()
            r2 = c2.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
            wo = _work_order_to_dict(c2, r2)
            c2.close()
            return jsonify({'work_order': wo, 'email_sent': False, 'email_error': None})
        if was_flagged:
            audit_desc = f"Part {idx+1} ({part_desc}) flag note updated — {note}"
            subj = f"[Part Flag Updated] Work Order — part note updated"
            body_extra = f"Part: {part_desc}\nStatus: PART FLAG UPDATED\n\nNew note:\n{note}\n"
            history_note = f"Part flag note updated — {part_desc}: {note}"
        else:
            audit_desc = f"Part {idx+1} ({part_desc}) flagged — {note}"
            subj = f"[Part Flagged] Work Order — part needs attention"
            body_extra = f"Part: {part_desc}\nStatus: PART FLAGGED\n\nReason:\n{note}\n"
            history_note = f"Part flagged — {part_desc}: {note}"
    else:
        updates = {'flagged': False, 'flag_note': ''}
        audit_desc = f"Part {idx+1} ({part_desc}) unflagged"
        subj = None
        body_extra = None
        history_note = f"Part unflagged — {part_desc}"
        history_note_type = 'general'

    wo, email_sent, email_error = _update_part_field(
        wid, idx, updates, audit_desc=audit_desc,
        email_subject=subj, email_body_extra=body_extra,
        history_note=history_note,
        history_note_type=history_note_type,
    )
    if wo is None:
        return jsonify({'error': 'Work order not found'}), 404
    return jsonify({'work_order': wo, 'email_sent': email_sent, 'email_error': email_error})


def _wo_part_lookup(conn, wid, part_key):
    """Return (work_order_row, part_dict) for the given WO + part_key, or (None, None)."""
    import json as json_mod
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        return None, None
    try:
        parts = json_mod.loads(row['parts_json'] or '[]') or []
    except Exception:
        parts = []
    for p in parts:
        if isinstance(p, dict) and str(p.get('key', '')) == part_key:
            return row, p
    return row, None


@app.route('/api/work-orders/<int:wid>/parts/<part_key>/photos', methods=['POST'])
@editor_required
def upload_work_order_part_photo(wid, part_key):
    """Upload one or more photos tied to a specific part of a work order.
    Accepts multipart/form-data with one or more `photo` fields and an
    optional `comment` field applied to each photo in this batch."""
    conn = get_db()
    wo_row, part = _wo_part_lookup(conn, wid, part_key)
    if not wo_row:
        conn.close()
        return jsonify({'error': 'Work order not found'}), 404
    if not part:
        conn.close()
        return jsonify({'error': 'Part not found'}), 404

    files = request.files.getlist('photo')
    if not files or not files[0].filename:
        files = request.files.getlist('photos')
    if not files or not files[0].filename:
        conn.close()
        return jsonify({'error': 'No photo provided'}), 400

    comment = str(request.form.get('comment', '') or '').strip()
    author = _actor()
    saved_ids = []
    saved_filenames = []
    for f in files:
        fname = save_image_resized(f)
        if not fname:
            continue
        cur = conn.execute(
            "INSERT INTO work_order_part_photos "
            "(work_order_id, part_key, filename, comment, uploaded_by) "
            "VALUES (?, ?, ?, ?, ?)",
            (wid, part_key, fname, comment, author)
        )
        saved_ids.append(cur.lastrowid)
        saved_filenames.append(fname)

    if not saved_ids:
        conn.close()
        return jsonify({'error': 'Could not process image'}), 400

    part_desc = f"{part.get('quantity', 1)} × {part.get('description', '')}"
    _log_wo_audit(conn, wid, 'edited',
                  f"Photo{'s' if len(saved_ids) > 1 else ''} added to part ({part_desc})"
                  + (f" — {comment}" if comment else ''))
    conn.execute(
        "INSERT INTO work_order_notes (work_order_id, note, author, note_type) VALUES (?, ?, ?, 'general')",
        (wid, f"Photo{'s' if len(saved_ids) > 1 else ''} added to part — {part_desc}"
              + (f": {comment}" if comment else ''),
         author)
    )
    conn.execute("UPDATE work_orders SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (wid,))
    conn.commit()

    # Email the salesperson with the photo(s) attached
    email_sent = False
    email_error = None
    sp_email = _lookup_salesperson_email(conn, wo_row['sales_person'])
    if sp_email:
        attachments = []
        for fname in saved_filenames:
            fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            try:
                with open(fpath, 'rb') as fh:
                    attachments.append({
                        'filename': fname,
                        'content': fh.read(),
                        'mime_type': 'image/jpeg',
                    })
            except Exception as e:
                app.logger.warning(f"Photo email attach read failed for {fname}: {e}")
        if attachments:
            count = len(attachments)
            subject = f"[Photo{'s' if count > 1 else ''}] Work Order {wo_row['wo_number']}"
            body_lines = [
                f"Work Order: {wo_row['wo_number']}",
                f"Customer: {wo_row['customer_name'] or '—'}",
                f"Vehicle: {wo_row['vehicle'] or '—'}",
                f"Part: {part_desc}",
                f"Uploaded by: {author}",
                f"{count} photo{'s' if count > 1 else ''} attached.",
            ]
            if comment:
                body_lines += ['', 'Comment:', comment]
            body = '\n'.join(body_lines) + '\n'
            email_sent, email_error = _send_email(sp_email, subject, body, attachments=attachments)
            desc = (f"Photo email sent to {sp_email}" if email_sent
                    else f"Photo email FAILED to {sp_email}: {email_error or 'unknown error'}")
            _log_wo_audit(conn, wid, 'edited', desc)
            conn.commit()

    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify({'work_order': wo, 'email_sent': email_sent, 'email_error': email_error})


@app.route('/api/work-orders/<int:wid>/photos/<int:photo_id>', methods=['PUT'])
@editor_required
def update_work_order_part_photo(wid, photo_id):
    """Update the comment on an existing photo."""
    data = request.get_json() or {}
    comment = str(data.get('comment', '') or '').strip()
    conn = get_db()
    photo = conn.execute(
        "SELECT id FROM work_order_part_photos WHERE id = ? AND work_order_id = ?",
        (photo_id, wid)
    ).fetchone()
    if not photo:
        conn.close()
        return jsonify({'error': 'Photo not found'}), 404
    conn.execute(
        "UPDATE work_order_part_photos SET comment = ? WHERE id = ?",
        (comment, photo_id)
    )
    _log_wo_audit(conn, wid, 'edited', f"Photo comment updated: {comment or '(empty)'}")
    conn.execute("UPDATE work_orders SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (wid,))
    conn.commit()
    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify(wo)


@app.route('/api/work-orders/<int:wid>/photos/<int:photo_id>', methods=['DELETE'])
@editor_required
def delete_work_order_part_photo(wid, photo_id):
    conn = get_db()
    photo = conn.execute(
        "SELECT id, filename, part_key FROM work_order_part_photos WHERE id = ? AND work_order_id = ?",
        (photo_id, wid)
    ).fetchone()
    if not photo:
        conn.close()
        return jsonify({'error': 'Photo not found'}), 404
    delete_image(photo['filename'])
    conn.execute("DELETE FROM work_order_part_photos WHERE id = ?", (photo_id,))
    _log_wo_audit(conn, wid, 'edited', "Photo removed from part")
    conn.execute("UPDATE work_orders SET updated_at = CURRENT_TIMESTAMP WHERE id = ?", (wid,))
    conn.commit()
    new_row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    wo = _work_order_to_dict(conn, new_row)
    conn.close()
    return jsonify(wo)


def _build_update_email_body(wo):
    """Compose a full-status plain-text email body for a work order."""
    status_label = {
        'requested': 'REQUESTED',
        'flagged': 'FLAGGED',
        'delivered': 'DELIVERED / COMPLETE',
    }.get(wo.get('status'), (wo.get('status') or '').upper())
    req_date = (wo.get('request_date') or '').split('.')[0].replace('T', ' ')
    comp_date = (wo.get('completed_at') or '').split('.')[0].replace('T', ' ') if wo.get('completed_at') else ''

    lines = []
    lines.append(f"Work Order: {wo.get('wo_number', '')}")
    lines.append(f"Status: {status_label}")
    if wo.get('priority'):
        lines.append(f"Priority: {wo.get('priority')}")
    lines.append('')
    lines.append(f"Customer: {wo.get('customer_name') or '—'}")
    lines.append(f"Sales Person: {wo.get('sales_person') or '—'}")
    lines.append(f"Warehouse: {wo.get('warehouse_location') or '—'}")
    lines.append(f"Vehicle: {wo.get('vehicle') or '—'}")
    lines.append(f"VIN: {wo.get('vin') or '—'}")
    if wo.get('quote_invoice'):
        lines.append(f"Quote/Invoice #: {wo.get('quote_invoice')}")
    lines.append(f"Requested: {req_date}")
    if comp_date:
        lines.append(f"Completed: {comp_date}")

    if (wo.get('notes') or '').strip():
        lines.append('')
        lines.append('Request Details:')
        lines.append(wo.get('notes').strip())

    if wo.get('status') == 'flagged' and (wo.get('flag_note') or '').strip():
        lines.append('')
        lines.append(f"Current flag reason: {wo.get('flag_note').strip()}")

    parts = wo.get('parts') or []
    if parts:
        lines.append('')
        lines.append('Parts Requested:')
        for p in parts:
            box = '[x]' if p.get('pulled') else '[ ]'
            flag = ' ⚑' if p.get('flagged') else ''
            row = f"  {box} {p.get('quantity', 1)} × {p.get('description', '')}{flag}"
            if p.get('pulled') and p.get('pulled_at'):
                row += f"  (pulled {p['pulled_at']})"
            if (p.get('details') or '').strip():
                row += f"\n       {p['details'].strip()}"
            if p.get('flagged') and p.get('flag_note'):
                row += f"\n       FLAGGED: {p['flag_note']}"
            lines.append(row)

    notes_hist = wo.get('notes_log') or []
    if notes_hist:
        lines.append('')
        lines.append('Notes & Activity:')
        for n in notes_hist:
            when = (n.get('created_at') or '').split('.')[0].replace('T', ' ')
            tag = '[FLAG] ' if n.get('note_type') == 'flag' else ''
            lines.append(f"  {when} — {n.get('author', '')}: {tag}{n.get('note', '')}")

    return '\n'.join(lines) + '\n'


@app.route('/api/work-orders/<int:wid>/send-update', methods=['POST'])
@editor_required
def send_work_order_update(wid):
    conn = get_db()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404

    wo = _work_order_to_dict(conn, row)
    sp_email = _lookup_salesperson_email(conn, row['sales_person'])
    if not sp_email:
        conn.close()
        return jsonify({'error': 'No email configured for this sales person'}), 400

    subject = f"[Update] Work Order {wo['wo_number']} — {wo.get('status', '').title()}"
    body = _build_update_email_body(wo)
    ok, err = _send_email(sp_email, subject, body)

    desc = f"Update email sent to {sp_email}" if ok else f"Update email FAILED to {sp_email}: {err or 'unknown error'}"
    _log_wo_audit(conn, wid, 'edited', desc)
    conn.commit()
    conn.close()

    if ok:
        return jsonify({'success': True, 'to': sp_email})
    return jsonify({'error': err or 'Send failed'}), 500


@app.route('/api/work-orders/<int:wid>/audit', methods=['GET'])
@audit_view_required
def get_work_order_audit(wid):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, action, actor, description, created_at FROM work_order_audit WHERE work_order_id = ? ORDER BY created_at DESC, id DESC",
        (wid,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/work-orders/<int:wid>/pdf')
@login_required
def work_order_pdf(wid):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas as pdf_canvas
    from io import BytesIO

    conn = get_db()
    row = conn.execute("SELECT * FROM work_orders WHERE id = ?", (wid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    wo = _work_order_to_dict(conn, row)
    conn.close()

    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=letter)
    w, h = letter

    # Header
    c.setFont("Helvetica-Bold", 20)
    c.drawString(0.6 * inch, h - 0.75 * inch, "Work Order")
    c.setFont("Helvetica-Bold", 14)
    c.drawRightString(w - 0.6 * inch, h - 0.75 * inch, wo['wo_number'])

    # Status badge text
    status_label = {
        'requested': 'REQUESTED',
        'flagged': 'FLAGGED',
        'delivered': 'DELIVERED / COMPLETE',
    }.get(wo['status'], wo['status'].upper())
    c.setFont("Helvetica", 10)
    c.drawRightString(w - 0.6 * inch, h - 1.0 * inch, f"Status: {status_label}")

    # Divider
    c.setStrokeColorRGB(0.7, 0.7, 0.7)
    c.line(0.6 * inch, h - 1.15 * inch, w - 0.6 * inch, h - 1.15 * inch)

    # Field table
    y = h - 1.5 * inch
    line_h = 0.26 * inch

    def _kv(label, value):
        nonlocal y
        c.setFont("Helvetica-Bold", 10)
        c.drawString(0.6 * inch, y, f"{label}:")
        c.setFont("Helvetica", 10)
        c.drawString(2.2 * inch, y, str(value or '—'))
        y -= line_h

    _kv("Request Date", (wo.get('request_date') or '').split('.')[0])
    _kv("Warehouse Location", wo.get('warehouse_location'))
    _kv("Customer Name", wo.get('customer_name'))
    _kv("Quote / Invoice #", wo.get('quote_invoice'))
    _kv("Sales Person", wo.get('sales_person'))
    _kv("Vehicle", wo.get('vehicle'))
    _kv("VIN", wo.get('vin'))
    _kv("Priority", wo.get('priority'))
    _kv("Created By", wo.get('created_by'))
    if wo.get('completed_at'):
        _kv("Completed", (wo.get('completed_at') or '').split('.')[0])

    # Parts requested
    parts = wo.get('parts') or []
    if parts:
        y -= line_h * 0.25
        c.setFont("Helvetica-Bold", 11)
        c.drawString(0.6 * inch, y, "Parts Requested")
        y -= 0.22 * inch
        c.setFont("Helvetica-Bold", 9)
        c.drawString(0.6 * inch, y, "Qty")
        c.drawString(1.2 * inch, y, "Description")
        y -= 0.18 * inch
        c.setFont("Helvetica", 10)
        for p in parts:
            if y < 1.2 * inch:
                c.showPage()
                y = h - 0.75 * inch
            c.drawString(0.6 * inch, y, str(p.get('quantity', 1)))
            desc_lines = _wrap_text(str(p.get('description', '')), 80)
            for i, line in enumerate(desc_lines):
                if y < 1.2 * inch:
                    c.showPage()
                    y = h - 0.75 * inch
                c.drawString(1.2 * inch, y, line)
                y -= 0.2 * inch
            # Short description rendered in italic beneath the main line
            details = (p.get('details') or '').strip()
            if details:
                c.setFont("Helvetica-Oblique", 9)
                for line in _wrap_text(details, 85):
                    if y < 1.2 * inch:
                        c.showPage()
                        y = h - 0.75 * inch
                    c.drawString(1.2 * inch, y, line)
                    y -= 0.18 * inch
                c.setFont("Helvetica", 10)

    # Request details block
    y -= line_h * 0.25
    c.setFont("Helvetica-Bold", 11)
    c.drawString(0.6 * inch, y, "Request Details")
    y -= 0.2 * inch
    c.setFont("Helvetica", 10)
    notes_text = (wo.get('notes') or '').strip() or '—'
    for line in _wrap_text(notes_text, 90):
        if y < 1.2 * inch:
            c.showPage()
            y = h - 0.75 * inch
        c.drawString(0.6 * inch, y, line)
        y -= 0.2 * inch

    # Activity & notes log
    if wo.get('notes_log'):
        y -= 0.1 * inch
        c.setFont("Helvetica-Bold", 11)
        c.drawString(0.6 * inch, y, "Notes & Activity")
        y -= 0.22 * inch
        c.setFont("Helvetica", 9)
        for n in wo['notes_log']:
            header = f"[{(n.get('created_at') or '').split('.')[0]}] {n.get('author', '')}"
            if y < 1.2 * inch:
                c.showPage()
                y = h - 0.75 * inch
            c.setFont("Helvetica-Bold", 9)
            c.drawString(0.6 * inch, y, header)
            y -= 0.18 * inch
            c.setFont("Helvetica", 9)
            for line in _wrap_text(n.get('note', ''), 100):
                if y < 1.2 * inch:
                    c.showPage()
                    y = h - 0.75 * inch
                c.drawString(0.75 * inch, y, line)
                y -= 0.18 * inch
            y -= 0.05 * inch

    # Footer
    c.setFont("Helvetica", 7)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawString(0.6 * inch, 0.5 * inch, f"Warehouse Manager v{APP_VERSION}")

    c.save()
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'inline; filename=work-order-{wo["wo_number"]}.pdf'}
    )


def _wrap_text(text, width):
    """Simple word-wrap to a character width. Preserves blank lines."""
    out = []
    for para in (text or '').split('\n'):
        if not para.strip():
            out.append('')
            continue
        words = para.split(' ')
        line = ''
        for w in words:
            if len(line) + len(w) + 1 <= width:
                line = (line + ' ' + w).strip()
            else:
                if line:
                    out.append(line)
                line = w
        if line:
            out.append(line)
    return out or ['']


# ══════════════════════════════════════════
#  HANDLE 401 FOR AJAX
# ══════════════════════════════════════════

@login_manager.unauthorized_handler
def unauthorized():
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.path.startswith('/api/'):
        return jsonify({'error': 'Authentication required'}), 401
    return redirect(url_for('login', next=request.url))


init_db()

if __name__ == '__main__':
    # Debug mode is opt-in via env to avoid shipping the Werkzeug debugger in
    # production (would expose a remote-code-execution console on errors).
    _debug = os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true', 'yes')
    app.run(host='0.0.0.0', port=5000, debug=_debug)
